from __future__ import annotations

import csv
import re
from collections import defaultdict
from copy import copy
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

from mesp_automation_engine import REPORTS, find_order, find_report, find_sample


TABLE_EXTENSIONS = {".xlsx", ".xlsm", ".csv"}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".pdf"}

BLUE = "1F4E78"
DARK_BLUE = "17365D"
LIGHT_BLUE = "D9EAF7"
LIGHT_GRAY = "F2F2F2"
WHITE = "FFFFFF"
BORDER = "B7C9D6"


def _safe(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _is_table(path: Path) -> bool:
    return path.suffix.lower() in TABLE_EXTENSIONS


def _is_image(path: Path) -> bool:
    return path.suffix.lower() in IMAGE_EXTENSIONS


def _kind(path: Path) -> str:
    return "table" if _is_table(path) else "image"


def _clean_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[:\\/?*\[\]]", "-", name)
    return cleaned[:31] or "Sheet"


def _unique_sheet_name(workbook: Workbook, base_name: str) -> str:
    base = _clean_sheet_name(base_name)
    if base not in workbook.sheetnames:
        return base
    index = 2
    while True:
        suffix = f"-{index}"
        candidate = _clean_sheet_name(base[: 31 - len(suffix)] + suffix)
        if candidate not in workbook.sheetnames:
            return candidate
        index += 1


def _style_title(cell) -> None:
    cell.font = Font(name="Arial", bold=True, size=14, color=DARK_BLUE)
    cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    cell.alignment = Alignment(vertical="center")


def _style_header(row) -> None:
    thin = Side(style="thin", color=BORDER)
    for cell in row:
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(name="Arial", bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)


def _style_body(sheet, min_row: int, max_row: int, max_col: int) -> None:
    thin = Side(style="thin", color=BORDER)
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)


def _auto_width(sheet, max_width: int = 42) -> None:
    widths: dict[int, int] = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value in (None, ""):
                continue
            length = min(max_width, max(10, len(str(cell.value)) + 2))
            widths[cell.column] = max(widths.get(cell.column, 10), length)
    for col, width in widths.items():
        sheet.column_dimensions[get_column_letter(col)].width = width


def _hyperlink(cell, target_sheet: str, display: str, target_cell: str = "A1") -> None:
    cell.value = display
    quoted_sheet = target_sheet.replace("'", "''")
    cell.hyperlink = Hyperlink(
        ref=cell.coordinate,
        location=f"'{quoted_sheet}'!{target_cell}",
        display=display,
    )
    cell.style = "Hyperlink"


def _discover_files(source_folder: Path) -> list[dict[str, Any]]:
    items = []
    for path in sorted(source_folder.rglob("*")):
        if not path.is_file():
            continue
        if path.name.startswith("~$"):
            continue
        if not (_is_table(path) or _is_image(path)):
            continue
        report = find_report(path.name)
        sample = find_sample(path.relative_to(source_folder))
        if not report or not sample:
            continue
        items.append(
            {
                "path": path,
                "relative_path": str(path.relative_to(source_folder)),
                "sample": sample,
                "order": find_order(path),
                "report": report,
                "kind": _kind(path),
            }
        )
    return items


def _sample_orders(items: list[dict[str, Any]]) -> dict[int, str]:
    orders: dict[int, str] = {}
    for item in items:
        order = item.get("order")
        if order and item["sample"] not in orders:
            orders[item["sample"]] = order
    return orders


def _copy_sheet_contents(source, target) -> None:
    target.sheet_view.showGridLines = False
    for merged_range in source.merged_cells.ranges:
        target.merge_cells(str(merged_range))

    for col_key, dimension in source.column_dimensions.items():
        target.column_dimensions[col_key].width = dimension.width

    for row_index, dimension in source.row_dimensions.items():
        target.row_dimensions[row_index].height = dimension.height

    for row in source.iter_rows():
        for source_cell in row:
            target_cell = target[source_cell.coordinate]
            target_cell.value = source_cell.value
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.fill = copy(source_cell.fill)
                target_cell.border = copy(source_cell.border)
                target_cell.alignment = copy(source_cell.alignment)
                target_cell.number_format = source_cell.number_format
                target_cell.protection = copy(source_cell.protection)
            if source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)
            if source_cell.comment:
                target_cell.comment = copy(source_cell.comment)


def _write_csv_sheet(path: Path, target) -> None:
    encodings = ["utf-8-sig", "gbk", "utf-8"]
    rows = None
    for encoding in encodings:
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                rows = list(csv.reader(handle))
            break
        except UnicodeDecodeError:
            continue
    if rows is None:
        rows = []

    for row_index, row in enumerate(rows, 1):
        for col_index, value in enumerate(row, 1):
            target.cell(row_index, col_index, value)

    if rows:
        _style_header(target[1])
        _style_body(target, 2, len(rows), max(len(row) for row in rows))
    _auto_width(target)


def _write_placeholder(sheet, title: str, source_path: Path) -> None:
    sheet["A1"] = title
    _style_title(sheet["A1"])
    sheet["A3"] = "源文件"
    sheet["B3"] = source_path.name
    sheet["A4"] = "说明"
    sheet["B4"] = "该文件未能复制为表格内容，请回到原始支持文件查看。"
    _auto_width(sheet)


def _create_table_sheet(workbook: Workbook, item: dict[str, Any], order: str) -> str:
    material_match = re.search(r"物料ID[-_\s]*([A-Za-z0-9]+)", item["path"].name, flags=re.IGNORECASE)
    if item.get("report") == "CKM3" and material_match:
        base_name = f"样本{item['sample']}.订单编号{order or '未识别'}-CKM3-{material_match.group(1)}"
    else:
        base_name = f"样本{item['sample']}.订单编号{order or '未识别'}-{item['report']}"
    name = _unique_sheet_name(
        workbook,
        base_name,
    )
    sheet = workbook.create_sheet(name)
    sheet.sheet_view.showGridLines = False
    path = item["path"]

    if path.suffix.lower() == ".csv":
        _write_csv_sheet(path, sheet)
    else:
        try:
            source_workbook = load_workbook(path, data_only=False)
            source_sheet = source_workbook[source_workbook.sheetnames[0]]
            _copy_sheet_contents(source_sheet, sheet)
            if sheet.max_row == 1 and sheet.max_column == 1 and sheet["A1"].value is None:
                _write_placeholder(sheet, name, path)
        except Exception:
            _write_placeholder(sheet, name, path)

    sheet.freeze_panes = "A2" if sheet.max_row > 1 else None
    if not any(sheet.column_dimensions[get_column_letter(col)].width for col in range(1, min(sheet.max_column, 20) + 1)):
        _auto_width(sheet)
    return name


def _build_directory_sheet(
    sheet,
    samples: list[int],
    orders: dict[int, str],
    table_links: dict[tuple[int, str], str],
    image_counts: dict[tuple[int, str], int],
    image_links: dict[tuple[int, str], str],
) -> None:
    sheet.title = "SPP目录"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A4"
    sheet["A1"] = "SPP Supporting Package 目录"
    _style_title(sheet["A1"])
    sheet.merge_cells("A1:J1")

    headers = ["样本", "订单编号"]
    for report in REPORTS:
        headers.extend([f"{report}-截图", f"{report}-表格"])
    for col, header in enumerate(headers, 1):
        sheet.cell(3, col, header)
    _style_header(sheet[3])

    for row_index, sample in enumerate(samples, 4):
        order = orders.get(sample, "")
        sheet.cell(row_index, 1, f"样本{sample}")
        sheet.cell(row_index, 2, order or "未识别")
        col = 3
        for report in REPORTS:
            count = image_counts.get((sample, report), 0)
            if count:
                _hyperlink(sheet.cell(row_index, col), "INF-截图", f"{count} 张", image_links.get((sample, report), "A1"))
            else:
                sheet.cell(row_index, col, "缺失")
            table_sheet = table_links.get((sample, report))
            if table_sheet:
                _hyperlink(sheet.cell(row_index, col + 1), table_sheet, "打开")
            else:
                sheet.cell(row_index, col + 1, "缺失")
            col += 2

    if samples:
        _style_body(sheet, 4, 3 + len(samples), len(headers))
    _auto_width(sheet)


def _fit_image(image: ExcelImage, max_width: int = 260, max_height: int = 165) -> None:
    width = image.width or max_width
    height = image.height or max_height
    scale = min(max_width / width, max_height / height, 1)
    image.width = int(width * scale)
    image.height = int(height * scale)


def _build_image_sheet(
    sheet,
    samples: list[int],
    orders: dict[int, str],
    image_items: list[dict[str, Any]],
) -> dict[tuple[int, str], str]:
    sheet.title = "INF-截图"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A3"
    sheet["A1"] = "INF-截图"
    _style_title(sheet["A1"])
    sheet.merge_cells("A1:E1")

    headers = ["样本 / 订单编号", *REPORTS]
    for col, header in enumerate(headers, 1):
        sheet.cell(3, col, header)
    _style_header(sheet[3])

    for col in range(1, 6):
        sheet.column_dimensions[get_column_letter(col)].width = 34 if col > 1 else 24

    images_by_sample_report: dict[tuple[int, str], list[dict[str, Any]]] = defaultdict(list)
    for item in image_items:
        images_by_sample_report[(item["sample"], item["report"])].append(item)

    row = 4
    image_links: dict[tuple[int, str], str] = {}
    for sample in samples:
        sheet.row_dimensions[row].height = 24
        sheet.cell(row, 1, f"样本{sample}\n订单编号：{orders.get(sample, '未识别')}")
        sheet.cell(row, 1).alignment = Alignment(vertical="top", wrap_text=True)
        block_height = 12
        for col_index, report in enumerate(REPORTS, 2):
            items = images_by_sample_report.get((sample, report), [])
            if not items:
                sheet.cell(row, col_index, "缺失")
                continue
            image_links[(sample, report)] = f"{get_column_letter(col_index)}{row}"
            cursor_row = row
            for item in items[:2]:
                path = item["path"]
                if path.suffix.lower() == ".pdf":
                    sheet.cell(cursor_row, col_index, f"PDF截图：{path.name}")
                    cursor_row += 2
                    continue
                try:
                    image = ExcelImage(path)
                    _fit_image(image)
                    image.anchor = f"{get_column_letter(col_index)}{cursor_row}"
                    sheet.add_image(image)
                    cursor_row += 9
                except Exception:
                    sheet.cell(cursor_row, col_index, path.name)
                    cursor_row += 2
            if len(items) > 2:
                sheet.cell(cursor_row, col_index, f"另有 {len(items) - 2} 个截图文件")
        for block_row in range(row, row + block_height):
            sheet.row_dimensions[block_row].height = 18
        row += block_height + 1

    if samples:
        _style_body(sheet, 4, row - 1, 5)

    return image_links


def build_supporting_workbook(result: dict, source_folder: Path) -> Workbook:
    items = _discover_files(source_folder)
    report_order = {report: index for index, report in enumerate(REPORTS)}
    items = sorted(
        items,
        key=lambda item: (
            item["sample"],
            report_order.get(item["report"], 99),
            0 if item["kind"] == "image" else 1,
            item["relative_path"],
        ),
    )
    samples = sorted({item["sample"] for item in items})
    orders = _sample_orders(items)

    workbook = Workbook()
    directory = workbook.active
    image_sheet = workbook.create_sheet("INF-截图")

    table_links: dict[tuple[int, str], str] = {}
    for item in items:
        if item["kind"] != "table":
            continue
        order = orders.get(item["sample"], item.get("order", ""))
        table_links[(item["sample"], item["report"])] = _create_table_sheet(workbook, item, order)

    image_items = [item for item in items if item["kind"] == "image"]
    image_counts = defaultdict(int)
    for item in image_items:
        image_counts[(item["sample"], item["report"])] += 1

    image_links = _build_image_sheet(image_sheet, samples, orders, image_items)
    _build_directory_sheet(directory, samples, orders, table_links, image_counts, image_links)

    workbook.properties.title = "SPP Supporting Package"
    workbook.properties.subject = result.get("program") or ""
    return workbook


def build_supporting_bytes(result: dict, source_folder: Path) -> bytes:
    workbook = build_supporting_workbook(result, source_folder)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()
