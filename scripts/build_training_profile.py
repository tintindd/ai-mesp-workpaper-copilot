from __future__ import annotations

import json
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image


WORKSPACE = Path(__file__).resolve().parents[1]
TRAINING_ROOT = WORKSPACE / "training_files" / "训练文件"
OUTPUT_PATH = WORKSPACE / "mesp_training_profile.json"

FILE_PATTERN = re.compile(
    r"(?P<sample>\d+)\.订单编号(?P<order>\d+)-(?P<report>[A-Za-z0-9]+)-(?P<kind>截图|表格)\.(?P<ext>\w+)$"
)

EXPECTED_EVIDENCE = [
    "CO03-截图",
    "CO03-表格",
    "KSBT-截图",
    "KSBT-表格",
    "3611-截图",
    "3611-表格",
    "CKM3-截图",
    "CKM3-表格",
]

FIELD_ALIASES = {
    "order_no": ["订单（原始）", "订单", "生产订单", "Order"],
    "cost_center": ["成本中心", "Cost Center"],
    "cost_element": ["成本要素", "Cost Element"],
    "cost_element_text": ["成本要素 (文本)", "成本要素文本", "Cost Element Text"],
    "material": ["物料", "Material"],
    "total_plan_cost": ["总计划成本", "Total Plan Cost"],
    "total_actual_cost": ["总实际成本", "Total Actual Cost"],
    "plan_actual_difference": ["计划/实际差异", "Plan/Actual Difference"],
    "actual_total_quantity": ["实际总计数量", "Actual Total Quantity"],
    "actual_fixed_cost": ["实际固定成本", "Actual Fixed Cost"],
    "actual_variable_cost": ["实际变动成本", "Actual Variable Cost"],
    "activity_type": ["活动类型", "Activity Type"],
}


def parse_file_name(path: Path) -> dict:
    match = FILE_PATTERN.match(path.name)
    parsed = {"file": path.name, "size": path.stat().st_size, "parsed": bool(match)}
    if match:
        parsed.update(
            {
                "sample": int(match.group("sample")),
                "order": match.group("order"),
                "report": match.group("report").upper(),
                "kind": match.group("kind"),
                "ext": match.group("ext").lower(),
            }
        )
    return parsed


def numeric(value) -> float:
    return value if isinstance(value, (int, float)) else 0.0


def inspect_workbook(path: Path) -> dict:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [value for value in header_row if value not in (None, "")]
    header_index = {value: index for index, value in enumerate(header_row)}

    totals = defaultdict(float)
    sample_values = defaultdict(list)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        for name in ["总计划成本", "总实际成本", "计划/实际差异", "实际总计数量"]:
            index = header_index.get(name)
            if index is not None and index < len(row):
                totals[name] += numeric(row[index])

        for name in ["成本中心", "成本要素", "成本要素 (文本)", "物料", "活动类型", "订单（原始）"]:
            index = header_index.get(name)
            if index is not None and index < len(row):
                value = row[index]
                if value not in (None, "") and len(sample_values[name]) < 8:
                    sample_values[name].append(value)

    matched_fields = {}
    missing_fields = []
    for canonical, aliases in FIELD_ALIASES.items():
        matched = next((alias for alias in aliases if alias in header_index), None)
        if matched:
            matched_fields[canonical] = {
                "source_header": matched,
                "column_index": header_index[matched] + 1,
            }
        else:
            missing_fields.append(canonical)

    return {
        "sheet": sheet.title,
        "rows": max((sheet.max_row or 1) - 1, 0),
        "cols": sheet.max_column,
        "headers": headers,
        "matched_fields": matched_fields,
        "missing_fields": missing_fields,
        "totals": {key: round(value, 2) for key, value in totals.items()},
        "sample_values": dict(sample_values),
    }


def inspect_image(path: Path) -> dict:
    image = Image.open(path)
    return {"width": image.width, "height": image.height, "mode": image.mode}


def build_profile() -> dict:
    files = []
    evidence_by_sample = defaultdict(set)
    co03_workbooks = []

    for path in sorted(TRAINING_ROOT.iterdir()):
        if not path.is_file():
            continue

        item = parse_file_name(path)
        if item.get("parsed"):
            evidence_by_sample[item["sample"]].add(f"{item['report']}-{item['kind']}")

        if path.suffix.lower() == ".xlsx":
            item["workbook"] = inspect_workbook(path)
            if item.get("report") == "CO03":
                co03_workbooks.append(
                    {
                        "file": item["file"],
                        "sample": item["sample"],
                        "order": item["order"],
                        **item["workbook"],
                    }
                )
        elif path.suffix.lower() in [".png", ".jpg", ".jpeg"]:
            item["image"] = inspect_image(path)

        files.append(item)

    sample_profiles = []
    for sample in sorted(evidence_by_sample):
        present = sorted(evidence_by_sample[sample])
        missing = [key for key in EXPECTED_EVIDENCE if key not in evidence_by_sample[sample]]
        sample_files = [file for file in files if file.get("sample") == sample]
        order = next((file.get("order") for file in sample_files if file.get("order")), None)
        sample_profiles.append(
            {
                "sample": sample,
                "order": order,
                "present_evidence": present,
                "missing_evidence": missing,
            }
        )

    common_co03_headers = []
    if co03_workbooks:
        header_sets = [set(book["headers"]) for book in co03_workbooks]
        common_co03_headers = sorted(set.intersection(*header_sets))

    return {
        "source_folder": str(TRAINING_ROOT),
        "training_summary": {
            "sample_count": len(sample_profiles),
            "file_count": len(files),
            "expected_evidence": EXPECTED_EVIDENCE,
            "known_limitation": "训练包中 KSBT、3611、CKM3 目前只有截图，没有对应表格；CO03 有可结构化 Excel 表格。",
        },
        "samples": sample_profiles,
        "co03_profile": {
            "workbook_count": len(co03_workbooks),
            "common_headers": common_co03_headers,
            "field_aliases": FIELD_ALIASES,
            "workbooks": co03_workbooks,
        },
        "files": files,
    }


def main() -> None:
    if not TRAINING_ROOT.exists():
        raise FileNotFoundError(f"Training folder not found: {TRAINING_ROOT}")

    profile = build_profile()
    OUTPUT_PATH.write_text(json.dumps(profile, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(f"Wrote {OUTPUT_PATH}")
    print(
        json.dumps(
            {
                "sample_count": profile["training_summary"]["sample_count"],
                "file_count": profile["training_summary"]["file_count"],
                "co03_workbook_count": profile["co03_profile"]["workbook_count"],
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
