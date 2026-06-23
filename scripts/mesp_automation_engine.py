from __future__ import annotations

import argparse
import csv
import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


REPORTS = ["CO03", "KSBT", "3611", "CKM3"]
TABLE_EXTENSIONS = {".xlsx", ".xlsm", ".csv"}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".pdf"}
EXPECTED_EVIDENCE = [f"{report}-{kind}" for report in REPORTS for kind in ["截图", "表格"]]

REQUIRED_FIELDS = {
    "CO03": {
        "cost_center": ["成本中心", "Cost Center"],
        "cost_element": ["成本要素", "Cost Element"],
        "cost_element_text": ["成本要素 (文本)", "成本要素文本", "Cost Element Text"],
        "material": ["物料", "Material"],
        "order_no": ["订单（原始）", "订单(原始)", "订单", "生产订单", "Order"],
        "total_plan_cost": ["总计划成本", "Total Plan Cost"],
        "total_actual_cost": ["总实际成本", "Total Actual Cost"],
        "plan_actual_difference": ["计划/实际差异", "Plan/Actual Difference"],
        "actual_total_quantity": ["实际总计数量", "Actual Total Quantity"],
        "actual_fixed_cost": ["实际固定成本", "Actual Fixed Cost"],
        "actual_variable_cost": ["实际变动成本", "实际可变成本", "Actual Variable Cost"],
        "activity_type": ["活动类型", "Activity Type"],
    },
    "KSBT": {
        "cost_center": ["成本中心", "Cost Center"],
        "activity_type": ["活动类型", "Activity Type"],
        "total_rate": ["Fix+可变价格", "固定+可变价格", "总价格", "Price"],
        "variable_rate": ["对象货币计价的变动成本", "变动成本", "Variable Cost"],
        "fixed_rate": ["对象货币计价的固定成本", "固定成本", "Fixed Cost"],
        "unit": ["单位", "Unit"],
        "currency": ["币种", "Currency"],
    },
    "3611": {
        "cost_element": ["成本要素编码", "成本要素", "Cost Element"],
        "cost_element_name": ["成本要素名称", "成本要素 (文本)", "Cost Element Name"],
        "actual_cost": ["实际成本", "Actual Cost"],
        "plan_cost": ["计划成本", "Plan Cost"],
        "absolute_difference": ["差异(绝对)", "差异", "Absolute Difference"],
        "difference_percent": ["差异(%)", "Difference %"],
    },
    "CKM3": {
        "category": ["类别", "Category"],
        "transaction_quantity": ["交易数量", "数量", "Transaction Quantity"],
        "unit": ["数量单位", "单位", "Unit"],
        "initial_valuation": ["初始评估", "期初金额", "Initial Valuation"],
        "price_difference": ["价格差异", "Price Difference"],
        "rate_difference": ["工率差异", "Rate Difference"],
        "actual_value": ["实际值", "Actual Value"],
        "direct_material": ["直接材料", "Direct Material"],
        "direct_labor": ["直接人工", "Direct Labor"],
        "indirect_labor": ["间接人工", "Indirect Labor"],
        "depreciation": ["折旧与摊销", "折旧", "Depreciation"],
        "energy": ["能耗", "Energy"],
        "other_manufacturing_overhead": ["其他制造费用", "Other Manufacturing Overhead"],
        "cost_component_total": ["成本构成总和", "总和", "Total"],
    },
}

TOTAL_FIELDS = {
    "CO03": ["总计划成本", "总实际成本", "计划/实际差异", "实际总计数量", "实际固定成本", "实际变动成本"],
    "KSBT": ["Fix+可变价格", "对象货币计价的变动成本", "对象货币计价的固定成本"],
    "3611": ["实际成本", "计划成本", "差异(绝对)"],
    "CKM3": ["交易数量", "初始评估", "价格差异", "工率差异", "实际值", "成本构成总和"],
}

SAMPLE_VALUE_FIELDS = {
    "CO03": ["成本中心", "成本要素", "成本要素 (文本)", "物料", "活动类型", "订单（原始）"],
    "KSBT": ["成本中心", "活动类型", "成本中心短文本", "作业类型短文本", "单位", "币种"],
    "3611": ["成本要素编码", "成本要素名称"],
    "CKM3": ["层级", "类别", "数量单位"],
}


def normalize_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    return re.sub(r"[\s_（）()\[\]【】\-—:：/]+", "", text)


def alias_index(values: Iterable) -> dict[str, int]:
    index = {}
    for i, value in enumerate(values):
        key = normalize_text(value)
        if key and key not in index:
            index[key] = i
    return index


def find_report(name: str) -> str | None:
    upper = name.upper()
    return next((report for report in REPORTS if report in upper), None)


def find_sample(path: Path) -> int | None:
    candidates = [path.stem, *[part for part in reversed(path.parts)]]
    patterns = [
        r"样本\s*(\d+)",
        r"sample\s*(\d+)",
        r"^(\d+)[.\-_ ]",
        r"^(\d+)$",
    ]
    for text in candidates:
        for pattern in patterns:
            match = re.search(pattern, text, flags=re.IGNORECASE)
            if match:
                sample = int(match.group(1))
                if 1 <= sample <= 999 and str(sample) not in REPORTS:
                    return sample
    return None


def find_order(path: Path) -> str:
    text = path.stem
    patterns = [
        r"订单编号\s*([A-Za-z0-9]+)",
        r"订单号\s*([A-Za-z0-9]+)",
        r"order\s*([A-Za-z0-9]+)",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            return match.group(1)

    long_number = re.search(r"(\d{6,})", text)
    return long_number.group(1) if long_number else ""


def find_kind(path: Path) -> str | None:
    suffix = path.suffix.lower()
    name = path.name.upper()
    if suffix in TABLE_EXTENSIONS or any(token in name for token in ["表格", "导出", "EXPORT", "EXCEL", "XLS", "CSV"]):
        return "表格"
    if suffix in IMAGE_EXTENSIONS or any(token in name for token in ["截图", "SCREEN", "IMAGE", "PNG", "JPG", "PDF"]):
        return "截图"
    return None


def parse_file(path: Path, root: Path) -> dict | None:
    report = find_report(path.name)
    sample = find_sample(path.relative_to(root))
    kind = find_kind(path)
    if not (report and sample and kind):
        return None

    return {
        "file": path.name,
        "relative_path": str(path.relative_to(root)),
        "path": str(path),
        "sample": sample,
        "order": find_order(path),
        "report": report,
        "kind": kind,
        "ext": path.suffix.lower().lstrip("."),
    }


def numeric(value) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.replace(",", "").replace("，", "").strip()
        if cleaned.startswith("(") and cleaned.endswith(")"):
            cleaned = "-" + cleaned[1:-1]
        try:
            return float(cleaned)
        except ValueError:
            return 0.0
    return 0.0


def detect_columns(report: str, headers: list) -> tuple[dict, list]:
    normalized_header_index = alias_index(headers)
    mapping = {}
    missing = []

    for canonical, aliases in REQUIRED_FIELDS.get(report, {}).items():
        matched_alias = None
        matched_index = None
        for alias in aliases:
            index = normalized_header_index.get(normalize_text(alias))
            if index is not None:
                matched_alias = headers[index]
                matched_index = index
                break
        if matched_index is not None:
            mapping[canonical] = {"source_header": matched_alias, "column_index": matched_index + 1}
        else:
            missing.append(canonical)

    return mapping, missing


def score_header_row(report: str, row: list) -> int:
    normalized = set(alias_index(row))
    score = 0
    for aliases in REQUIRED_FIELDS.get(report, {}).values():
        if any(normalize_text(alias) in normalized for alias in aliases):
            score += 1
    return score


def find_header_row(rows: list[list], report: str) -> tuple[int, list]:
    best_index = 0
    best_score = -1
    for index, row in enumerate(rows[:12]):
        score = score_header_row(report, row)
        if score > best_score:
            best_index = index
            best_score = score
    return best_index, rows[best_index] if rows else []


def load_tabular_rows(path: Path) -> tuple[str, list[list]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return path.stem, list(csv.reader(handle))

    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    return sheet.title, rows


def get_header_index(headers: list) -> dict[str, int]:
    return alias_index(headers)


def find_header(headers: list, header_name: str) -> int | None:
    return get_header_index(headers).get(normalize_text(header_name))


def analyze_workbook(path: Path, report: str) -> dict:
    sheet_name, rows = load_tabular_rows(path)
    if not rows:
        return {
            "report": report,
            "sheet": sheet_name,
            "row_count": 0,
            "column_count": 0,
            "mapping": {},
            "missing_fields": list(REQUIRED_FIELDS.get(report, {})),
            "totals": {},
            "sample_values": {},
        }

    header_row_index, headers = find_header_row(rows, report)
    mapping, missing_fields = detect_columns(report, headers)
    data_rows = rows[header_row_index + 1 :]

    totals = defaultdict(float)
    for row in data_rows:
        for header in TOTAL_FIELDS.get(report, []):
            index = find_header(headers, header)
            if index is not None and index < len(row):
                totals[header] += numeric(row[index])

    sample_values = defaultdict(list)
    for row in data_rows:
        for header in SAMPLE_VALUE_FIELDS.get(report, []):
            index = find_header(headers, header)
            if index is not None and index < len(row):
                value = row[index]
                if value not in (None, "") and len(sample_values[header]) < 8:
                    sample_values[header].append(value)

    return {
        "report": report,
        "sheet": sheet_name,
        "header_row": header_row_index + 1,
        "row_count": len([row for row in data_rows if any(value not in (None, "") for value in row)]),
        "column_count": len(headers),
        "headers": [value for value in headers if value not in (None, "")],
        "mapping": mapping,
        "missing_fields": missing_fields,
        "totals": {key: round(value, 2) for key, value in totals.items()},
        "sample_values": dict(sample_values),
    }


def build_issue(issue_type: str, sample, problem: str, suggestion: str, status: str = "warn") -> dict:
    return {"type": issue_type, "sample": sample, "problem": problem, "suggestion": suggestion, "status": status}


def build_trace_entries(item: dict, result: dict) -> list[dict]:
    entries = []
    for canonical, meta in result["mapping"].items():
        entries.append(
            {
                "sample": item["sample"],
                "order": item["order"],
                "report": item["report"],
                "workpaper_field": canonical,
                "source_file": item["file"],
                "source_sheet": result["sheet"],
                "source_header": meta["source_header"],
                "source_column": meta["column_index"],
                "status": "Mapped",
            }
        )
    return entries


def summarize_parse_failure(path: Path, root: Path) -> dict:
    return {
        "file": str(path.relative_to(root)),
        "sample_detected": find_sample(path.relative_to(root)),
        "report_detected": find_report(path.name),
        "kind_detected": find_kind(path),
        "reason": "需要能识别样本号、报表类型和文件类型",
    }


def analyze_folder(folder: Path, period: str = "", program: str = "") -> dict:
    parsed_files = []
    ignored_files = []

    for path in sorted(folder.rglob("*")):
        if not path.is_file():
            continue
        if path.name.startswith("~$"):
            continue
        parsed = parse_file(path, folder)
        if parsed:
            parsed_files.append(parsed)
        else:
            ignored_files.append(summarize_parse_failure(path, folder))

    by_sample = defaultdict(list)
    for item in parsed_files:
        by_sample[item["sample"]].append(item)

    issues = []
    workbook_results = []
    evidence_trace = []

    for sample in sorted(by_sample):
        sample_files = by_sample[sample]
        order = next((item["order"] for item in sample_files if item["order"]), "")
        present = {f"{item['report']}-{item['kind']}" for item in sample_files}

        for evidence_key in EXPECTED_EVIDENCE:
            if evidence_key not in present:
                report, kind = evidence_key.split("-")
                issues.append(
                    build_issue(
                        "支持文件缺失",
                        sample,
                        f"样本 {sample} 缺少 {report} {kind} 支持。",
                        f"请补充包含样本号、{report} 和 {kind} 标识的文件。",
                        "missing",
                    )
                )

        reports_seen = {item["report"] for item in sample_files}
        if len(reports_seen) < len(REPORTS):
            missing_reports = sorted(set(REPORTS) - reports_seen)
            issues.append(
                build_issue(
                    "报表类型缺失",
                    sample,
                    f"样本 {sample} 未识别到以下报表类型：{', '.join(missing_reports)}。",
                    "请确认文件名包含 CO03、KSBT、3611 或 CKM3。",
                    "missing",
                )
            )

        orders = {item["order"] for item in sample_files if item["order"]}
        if len(orders) > 1:
            issues.append(
                build_issue(
                    "订单编号不一致",
                    sample,
                    f"样本 {sample} 中识别到多个订单编号：{', '.join(sorted(orders))}。",
                    "请确认同一样本的截图和表格是否对应同一生产订单。",
                )
            )

        for item in sample_files:
            if item["kind"] == "表格" and item["ext"] in {"xlsx", "xlsm", "csv"}:
                result = analyze_workbook(Path(item["path"]), item["report"])
                result.update({"sample": sample, "order": order, "file": item["file"]})
                workbook_results.append(result)
                evidence_trace.extend(build_trace_entries(item, result))

                if result["missing_fields"]:
                    issues.append(
                        build_issue(
                            f"{item['report']} 字段缺失",
                            sample,
                            f"{item['report']} 表格缺少关键字段：{', '.join(result['missing_fields'])}。",
                            f"请在 SAP 导出 {item['report']} 时选择完整字段布局，或补充字段别名映射。",
                        )
                    )

    if period:
        issues.append(
            build_issue(
                "期间一致性检查",
                "全部",
                f"需要检查 KSBT、3611、CKM3 的会计期间是否与审计期间 {period} 一致。",
                "当前导出表格尚未统一出现会计期间字段；建议后续增加期间字段或在界面中让用户手工确认。",
            )
        )

    if program == "SPD03015" and any(item["report"] == "CKM3" for item in workbook_results):
        issues.append(
            build_issue(
                "SPD03015 计算口径待确认",
                "全部",
                "CKM3 表格已可结构化读取，但仍需确认哪些行分别代表期初、入库、出库和期末余额。",
                "建议下一步增加 CKM3 行类别映射规则，用于自动计算差异分摊率。",
            )
        )

    questions = [
        {"title": f"追问 {index + 1}: {issue['type']}", "body": issue["problem"], "ask": issue["suggestion"]}
        for index, issue in enumerate(issues)
    ]

    workbook_count_by_report = {
        report: sum(1 for item in workbook_results if item["report"] == report) for report in REPORTS
    }

    return {
        "input_folder": str(folder),
        "period": period,
        "program": program,
        "summary": {
            "sample_count": len(by_sample),
            "recognized_file_count": len(parsed_files),
            "ignored_file_count": len(ignored_files),
            "missing_file_count": sum(1 for issue in issues if issue["type"] in {"支持文件缺失", "报表类型缺失"}),
            "warning_count": sum(1 for issue in issues if issue["type"] not in {"支持文件缺失", "报表类型缺失"}),
            "workbook_count": len(workbook_results),
            "workbook_count_by_report": workbook_count_by_report,
            "co03_workbook_count": workbook_count_by_report.get("CO03", 0),
        },
        "ignored_files": ignored_files,
        "issues": issues,
        "questions": questions,
        "workbook_results": workbook_results,
        "co03_results": [item for item in workbook_results if item["report"] == "CO03"],
        "evidence_trace": evidence_trace,
    }


def main() -> None:
    workspace = Path(__file__).resolve().parents[1]
    parser = argparse.ArgumentParser(description="Analyze MESP support files and generate automation output.")
    parser.add_argument(
        "--input",
        default=str(workspace / "training_files" / "训练文件"),
        help="Folder containing MESP support files.",
    )
    parser.add_argument("--period", default="", help="Audit period, e.g. 2025.01.01-2025.12.31.")
    parser.add_argument("--program", default="", help="Selected MESP program, e.g. SPD03015.")
    parser.add_argument(
        "--out",
        default=str(workspace / "mesp_automation_result.json"),
        help="Output JSON path.",
    )
    args = parser.parse_args()

    result = analyze_folder(Path(args.input), period=args.period, program=args.program)
    output_path = Path(args.out)
    output_path.write_text(json.dumps(result, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(f"Wrote {output_path}")
    print(json.dumps(result["summary"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
