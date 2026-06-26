from __future__ import annotations

import re
from html.parser import HTMLParser
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


CKM3_HEADERS = [
    "层级",
    "类别",
    "交易数量",
    "数量单位",
    "初始评估",
    "价格差异",
    "工率差异",
    "实际值",
    "价格",
    "公司间利润",
    "直接材料",
    "直接人工",
    "间接人工",
    "折旧与摊销",
    "能耗",
    "低值易耗",
    "其他制造费用",
    "成本构成总和",
]


def build_ckm3_workbook_bytes(ocr_text: str) -> bytes:
    rows = extract_ckm3_rows(ocr_text)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"

    header_fill = PatternFill("solid", fgColor="00338D")
    header_font = Font(color="FFFFFF", bold=True)
    for col_index, header in enumerate(CKM3_HEADERS, start=1):
        cell = worksheet.cell(row=1, column=col_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_index, row in enumerate(rows, start=2):
        for col_index, header in enumerate(CKM3_HEADERS, start=1):
            worksheet.cell(row=row_index, column=col_index, value=row.get(header))

    worksheet.freeze_panes = "A2"
    for column in worksheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in column)
        worksheet.column_dimensions[column[0].column_letter].width = min(max_len + 2, 34)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def extract_ckm3_rows(ocr_text: str) -> list[dict[str, Any]]:
    parsed_rows = _extract_html_table_rows(ocr_text)
    if len(parsed_rows) < 2:
        return []

    rows = [_clean_ocr_row(row) for row in parsed_rows[1:]]
    rows = [row for row in rows if row]
    _adjust_contextual_signs(rows)
    levels = _infer_levels([str(row["类别"]) for row in rows])
    for row, level in zip(rows, levels):
        row["层级"] = level
    return [{header: row.get(header, 0 if header not in {"类别", "数量单位"} else "") for header in CKM3_HEADERS} for row in rows]


class _HTMLTableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.rows: list[list[str]] = []
        self._row: list[str] | None = None
        self._cell: list[str] | None = None
        self._in_cell = False

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag == "tr":
            self._row = []
        elif tag == "td":
            self._cell = []
            self._in_cell = True

    def handle_data(self, data: str) -> None:
        if self._in_cell and self._cell is not None:
            self._cell.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag == "td" and self._row is not None and self._cell is not None:
            self._row.append("".join(self._cell).strip())
            self._cell = None
            self._in_cell = False
        elif tag == "tr" and self._row is not None:
            self.rows.append(self._row)
            self._row = None


def _extract_html_table_rows(ocr_text: str) -> list[list[str]]:
    parser = _HTMLTableParser()
    parser.feed(ocr_text)
    return parser.rows


def _clean_ocr_row(raw_row: list[str]) -> dict[str, Any] | None:
    if len(raw_row) < 4:
        return None

    category = _normalize_category(raw_row[0])
    quantity, unit = _split_quantity(raw_row[1], raw_row[2] if len(raw_row) > 2 else "")
    numeric_start = 3
    if len(raw_row) > 2 and str(raw_row[2] or "").strip() and not re.search(r"[A-Za-z]", str(raw_row[2] or "")):
        numeric_start = 2
    numeric_values = [_parse_number(value) for value in raw_row[numeric_start:]]
    while len(numeric_values) < 14:
        numeric_values.append(0)
    numeric_values = numeric_values[:14]

    if numeric_values[-1] == 0 and numeric_values[-2] != 0:
        numeric_values[-1] = numeric_values[-2]
        numeric_values[-2] = 0
    if quantity < 0:
        for index in (0, 3, 6, 13):
            numeric_values[index] = -abs(numeric_values[index])
    if category == "结算" and numeric_values[3] > 0:
        for index in (3, 6, 13):
            numeric_values[index] = -abs(numeric_values[index])

    return {
        "类别": category,
        "交易数量": quantity,
        "数量单位": unit or "PC",
        "初始评估": numeric_values[0],
        "价格差异": numeric_values[1],
        "工率差异": numeric_values[2],
        "实际值": numeric_values[3],
        "价格": numeric_values[4],
        "公司间利润": numeric_values[5],
        "直接材料": numeric_values[6],
        "直接人工": numeric_values[7],
        "间接人工": numeric_values[8],
        "折旧与摊销": numeric_values[9],
        "能耗": numeric_values[10],
        "低值易耗": numeric_values[11],
        "其他制造费用": numeric_values[12],
        "成本构成总和": numeric_values[13],
    }


def _normalize_category(value: str) -> str:
    text = re.sub(r"\s+", " ", str(value or "").strip())
    text = re.sub(r"^(\d{10})(发票)", r"\1 \2", text)
    text = re.sub(r"^(\d{10})(有关订单的发货)", r"\1 \2", text)
    text = text.replace("余额（用途）", "杂额(用途)")
    text = text.replace("余额(用途)", "杂额(用途)")
    return text.replace("WIP 生产", "WIP生产")


def _split_quantity(quantity_text: str, unit_text: str) -> tuple[int | float, str]:
    text = f"{quantity_text}{unit_text}".replace(" ", "")
    match = re.search(r"(\d+-|-?\d+(?:\.\d+)?)\s*([A-Za-z]+)?", text)
    if not match:
        return 0, unit_text or "PC"
    number_text = match.group(1)
    if number_text.endswith("-"):
        number_text = f"-{number_text[:-1]}"
    number = float(number_text)
    return int(number) if number.is_integer() else number, match.group(2) or unit_text or "PC"


def _parse_number(value: str) -> int | float:
    text = str(value or "").strip().replace(",", "").replace("，", "")
    if not text:
        return 0
    if text.endswith("-"):
        text = f"-{text[:-1]}"
    try:
        number = float(text)
    except ValueError:
        return 0
    return int(number) if number.is_integer() else number


def _adjust_contextual_signs(rows: list[dict[str, Any]]) -> None:
    ending_inventory_delta: float | int | None = None
    in_consumption_section = False
    previous_category = ""
    for row in rows:
        category = str(row.get("类别") or "")
        initial_value = row.get("初始评估") or 0
        actual_value = row.get("实际值") or 0

        if category == "消耗":
            in_consumption_section = True
        elif category == "期末库存":
            in_consumption_section = False
            if previous_category != "期末库存":
                ending_inventory_delta = actual_value - initial_value

        if in_consumption_section and category in {"消耗", "杂额(用途)", "杂额（用途）", "余额（用途）", "WIP生产"}:
            row["价格差异"] = round(actual_value - initial_value, 2)
            if category in {"杂额(用途)", "杂额（用途）", "余额（用途）"} and row.get("交易数量"):
                row["价格"] = round(actual_value / row["交易数量"], 1)

        if category == "结算" and ending_inventory_delta is not None:
            sign = -1 if ending_inventory_delta < 0 else 1
            for key in ("实际值", "直接材料", "成本构成总和"):
                row[key] = sign * abs(row.get(key) or 0)
        previous_category = category


def _infer_levels(categories: list[str]) -> list[int]:
    levels: list[int] = []
    section = ""
    for index, category in enumerate(categories):
        if category in {"期初库存", "库存累计"}:
            section = category
            levels.append(0)
        elif category == "收货":
            section = "收货"
            levels.append(1)
        elif category == "采购订单":
            section = "采购订单"
            levels.append(2)
        elif "发票" in category:
            levels.append(3)
        elif category == "消耗":
            section = "消耗"
            levels.append(1)
        elif category == "生产":
            section = "生产"
            levels.append(2)
        elif category == "WIP生产":
            previous = categories[index - 1] if index > 0 else ""
            if previous == "生产":
                levels.append(2)
            else:
                section = "WIP生产"
                levels.append(1)
        elif category == "期末库存":
            previous = categories[index - 1] if index > 0 else ""
            section = "期末库存"
            levels.append(1 if previous == "期末库存" else 0)
        elif category == "结算":
            levels.append(1)
        elif category in {"杂额(用途)", "杂额（用途）", "余额（用途）", "生产", "WIP生产"} and section == "消耗":
            levels.append(2)
        elif "收货到库存" in category:
            levels.append(3)
        elif "SMKI" in category or "PRD" in category:
            levels.append(2 if section == "WIP生产" else 3)
        elif "有关订单的发货" in category:
            levels.append(4)
        elif "订单的在制品构建" in category:
            levels.append(4 if section == "生产" else 3)
        elif "在制品构建重估" in category:
            levels.append(3)
        else:
            levels.append(0)
    return levels
