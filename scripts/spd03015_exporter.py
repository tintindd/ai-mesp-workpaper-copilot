from __future__ import annotations

import re
from dataclasses import dataclass
from copy import copy
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, range_boundaries

from mesp_automation_engine import find_order, find_report, find_sample, numeric


BLUE = "1F4E78"
DARK_BLUE = "17365D"
LIGHT_BLUE = "D9EAF7"
LIGHT_YELLOW = "FFF2CC"
LIGHT_GRAY = "F2F2F2"
WHITE = "FFFFFF"
BORDER = "B7C9D6"
TEMPLATE_PATH = Path(__file__).resolve().parents[1] / "templates" / "spd03015_irm_sap_template.xlsx"
CKM3_LABEL_START_ROWS = {1: 10, 2: 56, 3: 117, 4: 167, 5: 213}
CKM3_TITLE_ROWS = {1: 1, 2: 49, 3: 99, 4: 148, 5: 197}
SPD03014_EXPENSES = [
    ("间接人工", "9043000200"),
    ("能耗", "9043000400"),
    ("低值易耗", "9043000500"),
    ("其他", "9043000600"),
]
SPD03014_TABLE_A_START_ROWS = {1: 23, 2: 47, 3: 71, 4: 95, 5: 119}
SPD03014_TABLE_B_START_ROW = 147
SPD03012_EXPENSES = [
    ("直接人工", "9043000100"),
    ("折旧与摊销", "9043000300"),
]
SPD03012_TABLE_A_START_ROWS = {1: 38, 2: 50, 3: 62, 4: 74, 5: 86}
SPD03012_TABLE_B_START_ROW = 103
IRM_EXPENSES = [*SPD03012_EXPENSES, *SPD03014_EXPENSES]
PROGRAM_SHEETS = {
    "SPD03012": "SPD03012-IRM(SAP)",
    "SPD03014": "SPD03014_IRM(SAP)",
    "SPD03015": "SPD03015_IRM(SAP)",
}


@dataclass
class Spd03015Sample:
    sample: int
    order: str
    material_id: str
    period: str
    ckm3_path: Path | None
    ckm3_rows: list[tuple[Any, ...]] | None
    beginning_qty: float
    beginning_variance: float
    receipt_qty: float
    receipt_variance: float
    outbound_qty: float
    outbound_variance: float


def _norm(value: Any) -> str:
    return str(value or "").strip()


def _material_id_from_name(name: str) -> str:
    match = re.search(r"物料ID[-_\s]*([A-Za-z0-9]+)", name, flags=re.IGNORECASE)
    return match.group(1) if match else ""


def _product_id_from_ckm3_rows(rows: list[tuple[Any, ...]] | None) -> str:
    for row in rows or []:
        text = " ".join(str(value) for value in row if value not in (None, ""))
        match = re.search(r"\b0{4,}(\d{8})\b", text)
        if match:
            return match.group(1)
    return ""


def _product_id_from_ckm3(path: Path | None) -> str:
    return _product_id_from_ckm3_rows(_load_first_sheet_rows(path))


def _safe_div(numerator: float, denominator: float) -> float:
    return numerator / denominator if denominator else 0.0


def _extract_ckm3_amounts_from_rows(rows: list[tuple[Any, ...]] | None) -> dict[str, float]:
    if not rows:
        return {}

    headers = [_norm(value) for value in rows[0]]
    index = {header: i for i, header in enumerate(headers) if header}

    def get(row: tuple, header: str) -> Any:
        position = index.get(header)
        if position is None or position >= len(row):
            return None
        return row[position]

    data_rows = rows[1:]

    def find_exact(category: str) -> tuple | None:
        return next((row for row in data_rows if _norm(get(row, "类别")) == category), None)

    def find_level_zero(category: str) -> tuple | None:
        return next(
            (
                row
                for row in data_rows
                if _norm(get(row, "类别")) == category and numeric(get(row, "层级")) == 0
            ),
            None,
        ) or find_exact(category)

    beginning = find_level_zero("期初库存")
    receipt = find_exact("收货")
    consumption = find_exact("消耗")
    ending = find_level_zero("期末库存")

    beginning_qty = numeric(get(beginning, "交易数量")) if beginning else 0.0
    beginning_variance = (
        numeric(get(beginning, "实际值")) - numeric(get(beginning, "初始评估")) if beginning else 0.0
    )

    if abs(beginning_variance) < 0.0000001:
        for row in data_rows:
            category = _norm(get(row, "类别"))
            if category in {"收货", "库存累计"}:
                break
            if "重新评估" in category or "上一期间结算" in category:
                beginning_variance += numeric(get(row, "价格差异"))

    receipt_qty = numeric(get(receipt, "交易数量")) if receipt else 0.0
    receipt_variance = numeric(get(receipt, "价格差异")) if receipt else 0.0

    outbound_row = consumption or ending
    outbound_qty = numeric(get(outbound_row, "交易数量")) if outbound_row else 0.0
    outbound_variance = numeric(get(outbound_row, "价格差异")) if outbound_row else 0.0

    return {
        "beginning_qty": beginning_qty,
        "beginning_variance": beginning_variance,
        "receipt_qty": receipt_qty,
        "receipt_variance": receipt_variance,
        "outbound_qty": outbound_qty,
        "outbound_variance": outbound_variance,
    }


def _extract_ckm3_amounts(path: Path | None) -> dict[str, float]:
    return _extract_ckm3_amounts_from_rows(_load_first_sheet_rows(path))


def _parse_report_sheet_name(sheet_name: str) -> tuple[int | None, str | None, str]:
    sample = find_sample(Path(sheet_name))
    report = find_report(sheet_name)
    order_match = re.search(r"订单编号\s*([A-Za-z0-9]+)", sheet_name, flags=re.IGNORECASE)
    order = order_match.group(1) if order_match else (find_order(Path(sheet_name)) or "")
    return sample, report, order


def _iter_candidate_workbooks(source_folder: Path) -> list[Path]:
    return [
        path
        for path in sorted(source_folder.rglob("*"))
        if path.is_file() and not path.name.startswith("~$") and path.suffix.lower() in {".xlsx", ".xlsm"}
    ]


def _load_spp_report_rows(source_folder: Path, sample: int, report: str) -> list[tuple[Any, ...]]:
    for workbook_path in _iter_candidate_workbooks(source_folder):
        try:
            workbook = load_workbook(workbook_path, data_only=True, read_only=True)
        except Exception:
            continue

        for sheet_name in workbook.sheetnames:
            sheet_sample, sheet_report, _ = _parse_report_sheet_name(sheet_name)
            if sheet_sample == sample and sheet_report == report:
                rows = list(workbook[sheet_name].iter_rows(values_only=True))
                workbook.close()
                return rows
        workbook.close()
    return []


def _spp_has_screenshot(source_folder: Path, sample: int, report: str) -> bool:
    report_columns = {
        "CO03": "CO03-截图",
        "KSBT": "KSBT-截图",
        "3611": "3611-截图",
        "CKM3": "CKM3-截图",
    }
    target_header = report_columns.get(report)
    if not target_header:
        return False

    for workbook_path in _iter_candidate_workbooks(source_folder):
        try:
            workbook = load_workbook(workbook_path, data_only=True, read_only=True)
        except Exception:
            continue
        if "SPP目录" not in workbook.sheetnames:
            workbook.close()
            continue

        sheet = workbook["SPP目录"]
        rows = list(sheet.iter_rows(values_only=True))
        workbook.close()
        if len(rows) < 4:
            continue

        headers = [_norm(value) for value in rows[2]]
        try:
            sample_col = headers.index("样本")
            report_col = headers.index(target_header)
        except ValueError:
            continue

        for row in rows[3:]:
            sample_text = _norm(row[sample_col] if sample_col < len(row) else "")
            if find_sample(Path(sample_text)) != sample:
                continue
            value = _norm(row[report_col] if report_col < len(row) else "")
            if value and value not in {"0", "0张", "0 张", "-"}:
                return True
    return False


def _load_report_rows(source_folder: Path, sample: int, report: str) -> list[tuple[Any, ...]]:
    report_path = _find_report_file(source_folder, sample, report)
    if report_path:
        return _load_first_sheet_rows(report_path)
    return _load_spp_report_rows(source_folder, sample, report)


def _discover_samples(source_folder: Path) -> list[Spd03015Sample]:
    by_sample: dict[int, dict[str, Any]] = {}

    for path in sorted(source_folder.rglob("*")):
        if not path.is_file() or path.name.startswith("~$"):
            continue

        sample = find_sample(path.relative_to(source_folder))
        if not sample:
            continue

        info = by_sample.setdefault(sample, {"sample": sample, "order": "", "material_id": "", "ckm3_path": None})
        order = find_order(path)
        material_id = _material_id_from_name(path.name)
        if order and not info["order"]:
            info["order"] = order
        if material_id and not info["material_id"]:
            info["material_id"] = material_id
        if find_report(path.name) == "CKM3" and path.suffix.lower() in {".xlsx", ".xlsm", ".csv"}:
            info["ckm3_path"] = path

    for workbook_path in _iter_candidate_workbooks(source_folder):
        try:
            workbook = load_workbook(workbook_path, data_only=True, read_only=True)
        except Exception:
            continue
        for sheet_name in workbook.sheetnames:
            sample, report, order = _parse_report_sheet_name(sheet_name)
            if not sample or report not in {"CO03", "KSBT", "3611", "CKM3"}:
                continue
            info = by_sample.setdefault(sample, {"sample": sample, "order": "", "material_id": "", "ckm3_path": None})
            if order and not info["order"]:
                info["order"] = order
            if report == "CKM3" and not info.get("ckm3_rows"):
                info["ckm3_rows"] = list(workbook[sheet_name].iter_rows(values_only=True))
        workbook.close()

    samples: list[Spd03015Sample] = []
    for sample_no in sorted(by_sample):
        info = by_sample[sample_no]
        ckm3_rows = info.get("ckm3_rows") or _load_spp_report_rows(source_folder, sample_no, "CKM3")
        amounts = (
            _extract_ckm3_amounts(info["ckm3_path"])
            if info.get("ckm3_path")
            else _extract_ckm3_amounts_from_rows(ckm3_rows)
        )
        samples.append(
            Spd03015Sample(
                sample=sample_no,
                order=info.get("order", ""),
                material_id=info.get("material_id", ""),
                period="",
                ckm3_path=info.get("ckm3_path"),
                ckm3_rows=ckm3_rows,
                beginning_qty=amounts.get("beginning_qty", 0.0),
                beginning_variance=amounts.get("beginning_variance", 0.0),
                receipt_qty=amounts.get("receipt_qty", 0.0),
                receipt_variance=amounts.get("receipt_variance", 0.0),
                outbound_qty=amounts.get("outbound_qty", 0.0),
                outbound_variance=amounts.get("outbound_variance", 0.0),
            )
        )

    return samples


def _style_header(sheet, row: int, max_col: int) -> None:
    thin = Side(style="thin", color=BORDER)
    for col in range(1, max_col + 1):
        cell = sheet.cell(row, col)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(name="Arial", bold=True, color=WHITE, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)


def _style_body(sheet, min_row: int, max_row: int, max_col: int) -> None:
    thin = Side(style="thin", color=BORDER)
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin, left=thin, right=thin)


def _set_widths(sheet) -> None:
    widths = {
        "A": 6,
        "B": 16,
        "C": 12,
        "D": 13,
        "E": 14,
        "F": 13,
        "G": 16,
        "H": 18,
        "I": 18,
        "J": 38,
        "K": 14,
        "L": 17,
        "M": 13,
        "N": 19,
        "O": 13,
        "P": 24,
        "Q": 24,
        "R": 24,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def _merge(sheet, ranges: list[str]) -> None:
    for cell_range in ranges:
        sheet.merge_cells(cell_range)


def _build_spd_sheet(workbook: Workbook, samples: list[Spd03015Sample], ck_rows: dict[int, dict[str, int]]) -> None:
    sheet = workbook.active
    sheet.title = "SPD03015_IRM(SAP)"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A19"
    _set_widths(sheet)

    _merge(sheet, ["A13:Q13", "A14:Q14", "B16:G16", "H16:I16", "K16:M16", "O16:Q16", "A28:Q28", "A30:R30"])

    sheet["A2"] = "kpmg    Inventory cost variance allocation (standard cost method) - recalculate\n存货成本差异分摊（标准成本法）- 重新计算"
    sheet["A3"] = "(This substantive procedure template is used to test the inventory cost variance allocation) (本实质性程序模板用于测试存货成本差异分摊)"
    sheet["A5"] = "Procedure ID 程序编号"
    sheet["C5"] = "SPD03015"
    sheet["A6"] = "Currency\n币种"
    sheet["C6"] = "Rmb\n人民币"
    sheet["A7"] = "Unit (e.g. 000s, millions)\n单位（例如千、百万）"
    sheet["C7"] = " "
    sheet["A9"] = "Period start date\n期间开始日"
    sheet["A10"] = "Period end date\n期间截止日"
    sheet["A13"] = "The testing steps are the steps listed for the procedure in the workflow screen. Add additional rows and columns as necessary if more testing steps are needed.\n测试步骤是指工作流程工作屏中列出的本程序相关步骤。如需增加测试步骤，请根据需要增加表格的行数和列数。"
    sheet["A14"] = "If the sample size for the substantive procedure is greater than the number of rows in the table below, copy and paste rows rather than inserting blank rows, to ensure that the formulas and pick-list options flow down to all rows.\n如果实质性程序的样本量超过下表行数，请进行复制粘贴（而非插入空白行）以确保公式和列表选项覆盖所有行。"

    sheet["B16"] = "Per production report/sub-ledger\n基于生产报告/明细账"
    sheet["H16"] = "Supporting document\n相关文件"
    sheet["K16"] = "4 = Reperform the calculation"
    sheet["O16"] = "Testing steps\n测试步骤"

    headers = [
        "序号",
        "物料 ID",
        "期间",
        "期初库存",
        "期初差异金额",
        "本期入库",
        "本月新增差异金额",
        "本月出库分摊的差异金额",
        "Working paper reference\n工作底稿索引",
        "[Other RDE, if applicable (please specify)\n其他相关数据要素，如适用（请注明）]",
        "The stage of completion\n完成阶段",
        "重新计算差异分摊率",
        "本月出库",
        "\n重新计算本月出库分摊的差异金额",
        "Variance\n差异",
        "评价期末存货差异计算方法的合理性。",
        "在根据完成阶段（即在制品、制成品）分配人工/间接制造费用的情况下，评估分配的合理性。",
        "3. Vouch the data used in the calculation to relevant documentation.\n将计算中使用的数据核对至相应的支持性文件。",
    ]
    for col, value in enumerate(headers, 1):
        sheet.cell(17, col, value)

    support_row = [
        "",
        "",
        "",
        "A-差异分摊表/CKM3",
        "B-差异分摊表/CKM3",
        "C-差异分摊表/CKM3",
        "D-差异分摊表/CKM3",
        "E-差异分摊表/CKM3",
        "",
        "",
        "",
        "F=（B+D)/(A+C)",
        "G-差异分摊表/CKM3",
        "H=F*G",
        "I=H-E",
        "",
        "",
        "",
    ]
    for col, value in enumerate(support_row, 1):
        sheet.cell(18, col, value)

    _style_header(sheet, 17, 18)
    for col in range(1, 19):
        sheet.cell(18, col).fill = PatternFill("solid", fgColor=LIGHT_YELLOW)
        sheet.cell(18, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    other_rde = (
        "【<差异分摊报表>表\n"
        "期初库存数量Inventory quantity at the beginning of the period\n"
        "本期入库数量Inventory receipt quantity in the current period\n"
        "期初库存差异Inventory amount variance at the beginning of the period\n"
        "本期入库差异Inventory receipt amount variance in the current period\n"
        "本期出库数量Inventory outbound quantity in the current period\n"
        "本期出库差异Inventory outbound amount variance in the current period\n\n"
        "【BOM】\n审批人\n审批时间"
    )

    for row_offset, sample in enumerate(samples):
        row = 19 + row_offset
        ck = ck_rows[sample.sample]
        sheet.cell(row, 1, sample.sample)
        sheet.cell(row, 2, sample.material_id or "")
        sheet.cell(row, 3, sample.period)
        sheet.cell(row, 4, f"='CKM3'!Z{ck['beginning_qty']}")
        sheet.cell(row, 5, f"='CKM3'!Z{ck['beginning_variance']}")
        sheet.cell(row, 6, f"='CKM3'!Z{ck['receipt_qty']}")
        sheet.cell(row, 7, f"='CKM3'!Z{ck['receipt_variance']}")
        sheet.cell(row, 8, f"='CKM3'!Z{ck['outbound_variance']}")
        sheet.cell(row, 9, "N/A")
        sheet.cell(row, 10, other_rde)
        sheet.cell(row, 12, f"=(E{row}+G{row})/(D{row}+F{row})")
        sheet.cell(row, 13, f"='CKM3'!Z{ck['outbound_qty']}")
        sheet.cell(row, 14, f"=M{row}*L{row}")
        sheet.cell(row, 15, f"=N{row}-H{row}")
        sheet.cell(row, 16, "Reasonable合理")
        sheet.cell(row, 17, "Reasonable合理")
        sheet.cell(row, 18, "Agrees相符")

    last_data_row = 18 + len(samples)
    if samples:
        _style_body(sheet, 18, last_data_row, 18)

    for row in range(19, last_data_row + 1):
        sheet.row_dimensions[row].height = 80
        for col in [4, 5, 6, 7, 8, 12, 13, 14, 15]:
            sheet.cell(row, col).number_format = "#,##0.00"

    sheet["A25"] = "Note: 项目组需根据被审计单位制定的存货差异分配方法修改重新计算的公式。"
    sheet["A27"] = "Note:\n注释："
    sheet["A28"] = "The design of this procedure template assumes that the relevant considerations for the procedure will be appropriately documented on the workflow screen for the procedure, unless they are directly addressed within this procedure template.\n本程序模板的设计基于以下假设：本程序的相关考虑事项将恰当记录于与本程序有关的工作流程工作屏中，除非它们已由本程序模板直接应对。 "
    sheet["A30"] = "The engagement team determines whether additional tick marks, other notations and/or additional information are relevant to document on this procedure template based on any custom steps that were added to the substantive procedure and/or the results obtained from performing the steps above.\n项目组应根据自主添加的任何实质性程序步骤和/或通过执行这些步骤获取的结果，确定是否需在本程序模板中添加额外的标记、其他注释及/或其他信息。 "

    for cell in ["A2", "A13", "A14", "A25", "A27", "A28", "A30"]:
        sheet[cell].alignment = Alignment(wrap_text=True, vertical="top")
    sheet["A2"].font = Font(name="Arial", bold=True, size=14, color=DARK_BLUE)
    for row in [13, 14, 28, 30]:
        sheet.cell(row, 1).fill = PatternFill("solid", fgColor=LIGHT_GRAY)


def _build_ckm3_sheet(workbook: Workbook, samples: list[Spd03015Sample]) -> dict[int, dict[str, int]]:
    sheet = workbook.create_sheet("CKM3")
    sheet.sheet_view.showGridLines = False
    sheet.column_dimensions["A"].width = 48
    sheet.column_dimensions["Y"].width = 24
    sheet.column_dimensions["Z"].width = 16

    ck_rows: dict[int, dict[str, int]] = {}
    for index, sample in enumerate(samples, 1):
        start = {1: 1, 2: 49, 3: 99, 4: 148, 5: 197}.get(index, 1 + (index - 1) * 49)
        sheet.cell(start, 1, f"{sample.sample}.订单编号：{sample.order or ''}")
        sheet.cell(start, 1).font = Font(name="Arial", bold=True, color=DARK_BLUE)

        labels = [
            ("beginning_qty", "期数库存交易数量", sample.beginning_qty),
            ("beginning_variance", "期初价格差异", sample.beginning_variance),
            ("receipt_qty", "收货库存交易数量", sample.receipt_qty),
            ("receipt_variance", "收货价格差异", sample.receipt_variance),
            ("outbound_qty", "消耗库存交易数量", sample.outbound_qty),
            ("outbound_variance", "消耗价格差异", sample.outbound_variance),
        ]

        label_start = start + 9
        ck_rows[sample.sample] = {}
        for offset, (key, label, value) in enumerate(labels):
            row = label_start + offset
            sheet.cell(row, 25, label)
            sheet.cell(row, 26, value)
            sheet.cell(row, 26).number_format = "#,##0.00"
            ck_rows[sample.sample][key] = row

    return ck_rows


def _ckm3_label_start(index: int) -> int:
    return CKM3_LABEL_START_ROWS.get(index, 10 + (index - 1) * 49)


def _ckm3_title_row(index: int) -> int:
    return CKM3_TITLE_ROWS.get(index, 1 + (index - 1) * 49)


def _populate_template_ckm3(sheet, samples: list[Spd03015Sample]) -> dict[int, dict[str, int]]:
    ck_rows: dict[int, dict[str, int]] = {}
    labels = [
        ("beginning_qty", "期数库存交易数量", "beginning_qty"),
        ("beginning_variance", "期初价格差异", "beginning_variance"),
        ("receipt_qty", "收货库存交易数量", "receipt_qty"),
        ("receipt_variance", "收货价格差异", "receipt_variance"),
        ("outbound_qty", "消耗库存交易数量", "outbound_qty"),
        ("outbound_variance", "消耗价格差异", "outbound_variance"),
    ]

    for index, sample in enumerate(samples, 1):
        title_row = _ckm3_title_row(index)
        label_start = _ckm3_label_start(index)
        sheet.cell(title_row, 1, f"{sample.sample}.订单编号：{sample.order or ''}")
        ck_rows[sample.sample] = {}

        for offset, (key, label, attr) in enumerate(labels):
            row = label_start + offset
            sheet.cell(row, 25, label)
            sheet.cell(row, 26, getattr(sample, attr))
            sheet.cell(row, 26).number_format = "#,##0.00"
            ck_rows[sample.sample][key] = row

    return ck_rows


def _copy_row_style(sheet, source_row: int, target_row: int, max_col: int = 37, copy_static_values: bool = False) -> None:
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height
    for col in range(1, max_col + 1):
        source = sheet.cell(source_row, col)
        target = sheet.cell(target_row, col)
        if source.has_style:
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.number_format = source.number_format
            target.protection = copy(source.protection)
        if copy_static_values and col in {9, 10, 16, 17, 18}:
            target.value = source.value


def _shift_merged_ranges_for_insert(sheet, insert_at: int, row_count: int) -> list[str]:
    merged_ranges_to_shift = [
        str(cell_range) for cell_range in sheet.merged_cells.ranges if cell_range.min_row >= insert_at
    ]
    for cell_range in merged_ranges_to_shift:
        sheet.unmerge_cells(cell_range)
    return merged_ranges_to_shift


def _restore_shifted_merged_ranges(sheet, merged_ranges: list[str], row_count: int) -> None:
    for cell_range in merged_ranges:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        shifted_range = (
            f"{get_column_letter(min_col)}{min_row + row_count}:"
            f"{get_column_letter(max_col)}{max_row + row_count}"
        )
        sheet.merge_cells(shifted_range)


def _delete_rows_preserving_merges(sheet, start_row: int, row_count: int) -> None:
    if row_count <= 0:
        return

    end_row = start_row + row_count - 1
    affected_ranges = [
        str(cell_range)
        for cell_range in sheet.merged_cells.ranges
        if cell_range.max_row >= start_row
    ]
    for cell_range in affected_ranges:
        sheet.unmerge_cells(cell_range)

    sheet.delete_rows(start_row, row_count)

    for cell_range in affected_ranges:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        if min_row >= start_row and max_row <= end_row:
            continue
        if min_row > end_row:
            min_row -= row_count
            max_row -= row_count
        elif min_row < start_row <= max_row:
            max_row = max(start_row - 1, max_row - row_count)
        elif start_row <= min_row <= end_row < max_row:
            min_row = start_row
            max_row -= row_count
        if min_row <= max_row:
            sheet.merge_cells(
                f"{get_column_letter(min_col)}{min_row}:"
                f"{get_column_letter(max_col)}{max_row}"
            )


def _ensure_template_sample_rows(sheet, sample_count: int) -> None:
    if sample_count < 5:
        _delete_rows_preserving_merges(sheet, 19 + sample_count, 5 - sample_count)
        return
    if sample_count == 5:
        return

    extra_rows = sample_count - 5
    insert_at = 24
    template_row = 23
    static_values = {col: sheet.cell(template_row, col).value for col in [9, 10, 16, 17, 18]}
    merged_ranges_to_shift = _shift_merged_ranges_for_insert(sheet, insert_at, extra_rows)

    sheet.insert_rows(insert_at, extra_rows)
    _restore_shifted_merged_ranges(sheet, merged_ranges_to_shift, extra_rows)

    for row in range(insert_at, insert_at + extra_rows):
        _copy_row_style(sheet, template_row, row, copy_static_values=False)
        for col, value in static_values.items():
            sheet.cell(row, col, value)


def _populate_template_spd(sheet, samples: list[Spd03015Sample]) -> None:
    blue_fill = PatternFill("solid", fgColor="FF00338D")
    for row in [2, 3]:
        for col in range(1, 38):
            sheet.cell(row, col).fill = blue_fill

    _ensure_template_sample_rows(sheet, len(samples))

    last_sample_row = 18 + len(samples)
    for row in range(19, last_sample_row + 1):
        for col in [2, 3, 4, 5, 6, 7, 8, 12, 13, 14, 15]:
            sheet.cell(row, col).value = None

    for row_offset, sample in enumerate(samples):
        row = 19 + row_offset
        sheet.cell(row, 1, sample.sample)
        sheet.cell(row, 2, sample.material_id or "")
        sheet.cell(row, 3, sample.period or "")
        sheet.cell(row, 4, sample.beginning_qty)
        sheet.cell(row, 5, sample.beginning_variance)
        sheet.cell(row, 6, sample.receipt_qty)
        sheet.cell(row, 7, sample.receipt_variance)
        sheet.cell(row, 8, sample.outbound_variance)
        sheet.cell(row, 12, f"=(E{row}+G{row})/(D{row}+F{row})")
        sheet.cell(row, 13, sample.outbound_qty)
        sheet.cell(row, 14, f"=M{row}*L{row}")
        sheet.cell(row, 15, f"=N{row}-H{row}")


def _header_index(headers: tuple[Any, ...]) -> dict[str, int]:
    return {_norm(header): index for index, header in enumerate(headers) if _norm(header)}


def _cell_value(row: tuple[Any, ...], index: dict[str, int], *names: str) -> Any:
    for name in names:
        position = index.get(name)
        if position is not None and position < len(row):
            return row[position]
    return None


def _find_report_file(source_folder: Path, sample: int, report: str) -> Path | None:
    for path in sorted(source_folder.rglob("*")):
        if not path.is_file() or path.name.startswith("~$"):
            continue
        if find_sample(path.relative_to(source_folder)) == sample and find_report(path.name) == report:
            if path.suffix.lower() in {".xlsx", ".xlsm"}:
                return path
    return None


def _document_type_for_report(source_folder: Path, sample: int, report: str, prefer_image: bool) -> str:
    image_exts = {".png", ".jpg", ".jpeg", ".pdf"}
    table_exts = {".xlsx", ".xlsm", ".csv"}
    matched_exts = []
    for path in sorted(source_folder.rglob("*")):
        if not path.is_file() or path.name.startswith("~$"):
            continue
        if find_sample(path.relative_to(source_folder)) == sample and find_report(path.name) == report:
            matched_exts.append(path.suffix.lower())

    if prefer_image and any(ext in image_exts for ext in matched_exts):
        return "截屏类电子文件【Electronic document】"
    if any(ext in table_exts for ext in matched_exts):
        return "清单类电子数据【Electronic data】"
    if any(ext in image_exts for ext in matched_exts):
        return "截屏类电子文件【Electronic document】"
    if prefer_image and _spp_has_screenshot(source_folder, sample, report):
        return "截屏类电子文件【Electronic document】"
    if _load_spp_report_rows(source_folder, sample, report):
        return "清单类电子数据【Electronic data】"
    return ""


def _load_first_sheet_rows(path: Path | None) -> list[tuple[Any, ...]]:
    if not path:
        return []
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    return rows


def _coerce_rows(source: Path | list[tuple[Any, ...]] | None) -> list[tuple[Any, ...]]:
    if isinstance(source, list):
        return source
    return _load_first_sheet_rows(source)


def _extract_co03_spd03014(source: Path | list[tuple[Any, ...]] | None) -> dict[str, Any]:
    rows = _coerce_rows(source)
    if not rows:
        return {"product_id": "", "cost_center": "", "expenses": {}}

    index = _header_index(rows[0])
    product_id = ""
    cost_center = ""
    expenses: dict[str, dict[str, float]] = {}

    for row in rows[1:]:
        business = str(_cell_value(row, index, "业务交易") or "")
        material = _cell_value(row, index, "物料", "来源")
        if not product_id and "收货" in business and material not in (None, ""):
            match = re.search(r"(\d{8})", str(material))
            if match:
                product_id = match.group(1)

        cost_element = str(_cell_value(row, index, "成本要素") or "")
        for expense, code in IRM_EXPENSES:
            if cost_element == code:
                cost_center = str(_cell_value(row, index, "成本中心") or cost_center)
                expenses[expense] = {
                    "qty": numeric(_cell_value(row, index, "实际总计数量")),
                    "plan": numeric(_cell_value(row, index, "总计划成本")),
                    "actual": numeric(_cell_value(row, index, "总实际成本")),
                    "variance": numeric(_cell_value(row, index, "计划/实际差异")),
                }

    return {"product_id": product_id, "cost_center": cost_center, "expenses": expenses}


def _extract_ksbt_rates(source: Path | list[tuple[Any, ...]] | None) -> dict[str, dict[str, float]]:
    rows = _coerce_rows(source)
    if not rows:
        return {}
    index = _header_index(rows[0])
    rates: dict[str, dict[str, float]] = {expense: {"plan": 0.0, "actual": 0.0} for expense, _ in IRM_EXPENSES}

    for row in rows[1:]:
        text = str(_cell_value(row, index, "作业类型短文本") or "")
        price = abs(numeric(_cell_value(row, index, "Fix+可变价格", "固定+可变价格", "总价格")))
        row_type = str(_cell_value(row, index, "A") or "").upper()
        prt = numeric(_cell_value(row, index, "PrT"))
        for expense, _ in IRM_EXPENSES:
            if expense in text:
                if row_type == "A" or prt == 5:
                    rates[expense]["actual"] = price
                elif not rates[expense]["plan"]:
                    rates[expense]["plan"] = price

    return rates


def _extract_3611_amounts(source: Path | list[tuple[Any, ...]] | None) -> dict[str, float]:
    rows = _coerce_rows(source)
    if not rows:
        return {}
    index = _header_index(rows[0])
    amounts = {expense: 0.0 for expense, _ in IRM_EXPENSES}

    for row in rows[1:]:
        name = str(_cell_value(row, index, "成本要素名称", "成本要素 (文本)") or "")
        amount = abs(numeric(_cell_value(row, index, "实际成本", "差异(绝对)")))
        for expense, _ in IRM_EXPENSES:
            if expense in name or (expense == "其他" and "其他" in name):
                amounts[expense] = amount

    return amounts


def _build_spd03014_data(source_folder: Path, samples: list[Spd03015Sample]) -> list[dict[str, Any]]:
    results = []
    for sample in samples:
        co03 = _extract_co03_spd03014(_load_report_rows(source_folder, sample.sample, "CO03"))
        ksbt = _extract_ksbt_rates(_load_report_rows(source_folder, sample.sample, "KSBT"))
        amounts_3611 = _extract_3611_amounts(_load_report_rows(source_folder, sample.sample, "3611"))
        product_id = (
            co03.get("product_id")
            or _product_id_from_ckm3_rows(sample.ckm3_rows)
            or _product_id_from_ckm3(sample.ckm3_path)
            or sample.material_id
            or sample.order
        )
        energy_rate = (ksbt.get("能耗") or {}).get("actual", 0.0)
        common_overhead_hours = (
            round(amounts_3611.get("能耗", 0.0) / energy_rate) if energy_rate else 0
        )

        expense_rows = []
        for expense, _ in SPD03014_EXPENSES:
            co03_expense = (co03.get("expenses") or {}).get(expense, {})
            actual_rate = (ksbt.get(expense) or {}).get("actual", 0.0)
            plan_rate = (ksbt.get(expense) or {}).get("plan", 0.0)
            amount_3611 = amounts_3611.get(expense, 0.0)
            denominator = common_overhead_hours or (round(amount_3611 / actual_rate) if actual_rate else 0)
            actual_absorption_rate = amount_3611 / denominator if denominator else 0.0

            expense_rows.append(
                {
                    "expense": expense,
                    "qty": co03_expense.get("qty", 0.0),
                    "plan": co03_expense.get("plan", 0.0),
                    "actual": co03_expense.get("actual", 0.0),
                    "variance": co03_expense.get("variance", 0.0),
                    "actual_rate": actual_rate,
                    "plan_rate": plan_rate,
                    "actual_absorption_rate": actual_absorption_rate,
                }
            )

        results.append(
            {
                "sample": sample.sample,
                "order": sample.order,
                "product_id": product_id,
                "cost_center": "",
                "expenses": expense_rows,
            }
        )

    return results


def _copy_row_all(sheet, source_row: int, target_row: int, max_col: int = 26) -> None:
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height
    for col in range(1, max_col + 1):
        source = sheet.cell(source_row, col)
        target = sheet.cell(target_row, col)
        target.value = source.value
        if source.has_style:
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.number_format = source.number_format
            target.protection = copy(source.protection)


def _copy_merged_ranges_with_offset(sheet, source_start: int, source_end: int, row_offset: int) -> None:
    source_ranges = [
        str(cell_range)
        for cell_range in sheet.merged_cells.ranges
        if cell_range.min_row >= source_start and cell_range.max_row <= source_end
    ]
    for cell_range in source_ranges:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        shifted_range = (
            f"{get_column_letter(min_col)}{min_row + row_offset}:"
            f"{get_column_letter(max_col)}{max_row + row_offset}"
        )
        sheet.merge_cells(shifted_range)


def _ensure_spd03014_sample_rows(sheet, sample_count: int) -> None:
    if sample_count < 5:
        table_a_delete_start = 23 + sample_count * 24
        table_a_rows = (5 - sample_count) * 24
        _delete_rows_preserving_merges(sheet, table_a_delete_start, table_a_rows)

        table_b_start = SPD03014_TABLE_B_START_ROW - table_a_rows
        table_b_delete_start = table_b_start + sample_count * 4
        _delete_rows_preserving_merges(sheet, table_b_delete_start, (5 - sample_count) * 4)
        return
    if sample_count == 5:
        return

    extra_samples = sample_count - 5
    table_a_insert_at = 143
    table_a_rows = extra_samples * 24
    shifted = _shift_merged_ranges_for_insert(sheet, table_a_insert_at, table_a_rows)
    sheet.insert_rows(table_a_insert_at, table_a_rows)
    _restore_shifted_merged_ranges(sheet, shifted, table_a_rows)

    for extra_index in range(extra_samples):
        target_start = table_a_insert_at + extra_index * 24
        row_offset = target_start - 119
        for offset in range(24):
            _copy_row_all(sheet, 119 + offset, target_start + offset)
        _copy_merged_ranges_with_offset(sheet, 119, 142, row_offset)

    table_b_insert_at = 167 + table_a_rows
    table_b_rows = extra_samples * 4
    shifted = _shift_merged_ranges_for_insert(sheet, table_b_insert_at, table_b_rows)
    sheet.insert_rows(table_b_insert_at, table_b_rows)
    _restore_shifted_merged_ranges(sheet, shifted, table_b_rows)

    for extra_index in range(extra_samples):
        target_start = table_b_insert_at + extra_index * 4
        for offset in range(4):
            _copy_row_all(sheet, 163 + offset + table_a_rows, target_start + offset)


def _spd03014_table_a_start(sample_index: int) -> int:
    if sample_index <= 5:
        return SPD03014_TABLE_A_START_ROWS[sample_index]
    return 143 + (sample_index - 6) * 24


def _spd03014_table_b_start(sample_index: int, sample_count: int) -> int:
    if sample_count < 5:
        base = SPD03014_TABLE_B_START_ROW - (5 - sample_count) * 24
    else:
        base = SPD03014_TABLE_B_START_ROW + max(sample_count - 5, 0) * 24
    return base + (sample_index - 1) * 4


def _populate_template_spd03014(sheet, source_folder: Path, samples: list[Spd03015Sample]) -> None:
    _ensure_spd03014_sample_rows(sheet, len(samples))
    data = _build_spd03014_data(source_folder, samples)
    for sample_index, sample_data in enumerate(data, 1):
        block_start = _spd03014_table_a_start(sample_index)
        sample_no = sample_data["sample"]
        screenshot_type = _document_type_for_report(source_folder, sample_no, "CO03", prefer_image=True)
        co03_table_type = _document_type_for_report(source_folder, sample_no, "CO03", prefer_image=False)

        for expense_index, expense_data in enumerate(sample_data["expenses"]):
            row = block_start + expense_index * 6
            sheet.cell(row, 2, f"样本{sample_data['sample']}")
            sheet.cell(row, 3, sample_data["product_id"])
            sheet.cell(row, 4, sample_data["cost_center"])
            sheet.cell(row, 5, sample_data["order"])
            sheet.cell(row, 6, expense_data["expense"])
            sheet.cell(row, 7, expense_data["qty"])
            sheet.cell(row, 8, expense_data["plan"])
            sheet.cell(row, 9, expense_data["actual"])
            sheet.cell(row, 10, expense_data["variance"])
            sheet.cell(row, 11, expense_data["actual_rate"])
            sheet.cell(row, 12, f"=IF(N{row}=0,0*G{row}-H{row},I{row}/N{row}*G{row}-H{row})")
            sheet.cell(row, 13, f"=L{row}-J{row}")
            sheet.cell(row, 14, f"=G{row}")
            sheet.cell(row, 15, "不适用")
            sheet.cell(row, 16, expense_data["actual_absorption_rate"])
            sheet.cell(row, 17, f"=P{row}-K{row}")
            sheet.cell(row, 20, f"=N{row}")
            for doc_offset in range(6):
                sheet.cell(row + doc_offset, 19, co03_table_type if doc_offset == 4 else screenshot_type)

        table_b_start = _spd03014_table_b_start(sample_index, len(data))
        for expense_index, expense_data in enumerate(sample_data["expenses"]):
            table_a_row = block_start + expense_index * 6
            row = table_b_start + expense_index
            qty = expense_data["qty"]
            standard_unit_cost = expense_data["plan"] / qty if qty else expense_data["plan_rate"]
            standard_total = standard_unit_cost * qty if qty else expense_data["plan"]

            sheet.cell(row, 2, f"样本{sample_data['sample']}")
            sheet.cell(row, 3, sample_data["product_id"])
            sheet.cell(row, 4, sample_data["cost_center"])
            sheet.cell(row, 5, sample_data["order"])
            sheet.cell(row, 6, expense_data["expense"])
            sheet.cell(row, 7, standard_unit_cost)
            sheet.cell(row, 8, qty)
            sheet.cell(row, 9, standard_total)
            sheet.cell(row, 10, expense_data["actual"])
            sheet.cell(row, 11, f"=J{row}-I{row}")
            sheet.cell(row, 12, f"=J{table_a_row}")
            sheet.cell(row, 13, f"=L{row}-K{row}")


def _build_spd03012_data(source_folder: Path, samples: list[Spd03015Sample]) -> list[dict[str, Any]]:
    results = []
    for sample in samples:
        co03 = _extract_co03_spd03014(_load_report_rows(source_folder, sample.sample, "CO03"))
        ksbt = _extract_ksbt_rates(_load_report_rows(source_folder, sample.sample, "KSBT"))
        amounts_3611 = _extract_3611_amounts(_load_report_rows(source_folder, sample.sample, "3611"))
        product_id = (
            co03.get("product_id")
            or _product_id_from_ckm3_rows(sample.ckm3_rows)
            or _product_id_from_ckm3(sample.ckm3_path)
            or sample.material_id
            or sample.order
        )

        expense_rows = []
        for expense, _ in SPD03012_EXPENSES:
            co03_expense = (co03.get("expenses") or {}).get(expense, {})
            actual_rate = (ksbt.get(expense) or {}).get("actual", 0.0)
            amount_3611 = amounts_3611.get(expense, 0.0)
            denominator = round(amount_3611 / actual_rate) if actual_rate else 0
            actual_absorption_rate = amount_3611 / denominator if denominator else 0.0
            expense_rows.append(
                {
                    "expense": expense,
                    "qty": co03_expense.get("qty", 0.0),
                    "plan": co03_expense.get("plan", 0.0),
                    "actual": co03_expense.get("actual", 0.0),
                    "variance": co03_expense.get("variance", 0.0),
                    "actual_rate": actual_rate,
                    "plan_rate": (ksbt.get(expense) or {}).get("plan", 0.0),
                    "actual_absorption_rate": actual_absorption_rate,
                }
            )

        results.append(
            {
                "sample": sample.sample,
                "order": sample.order,
                "product_id": product_id,
                "cost_center": "",
                "expenses": expense_rows,
            }
        )

    return results


def _ensure_spd03012_sample_rows(sheet, sample_count: int) -> None:
    if sample_count < 5:
        table_a_delete_start = 38 + sample_count * 12
        table_a_rows = (5 - sample_count) * 12
        _delete_rows_preserving_merges(sheet, table_a_delete_start, table_a_rows)

        table_b_start = SPD03012_TABLE_B_START_ROW - table_a_rows
        table_b_delete_start = table_b_start + sample_count * 2
        _delete_rows_preserving_merges(sheet, table_b_delete_start, (5 - sample_count) * 2)
        return
    if sample_count == 5:
        return

    extra_samples = sample_count - 5
    table_a_insert_at = 98
    table_a_rows = extra_samples * 12
    shifted = _shift_merged_ranges_for_insert(sheet, table_a_insert_at, table_a_rows)
    sheet.insert_rows(table_a_insert_at, table_a_rows)
    _restore_shifted_merged_ranges(sheet, shifted, table_a_rows)

    for extra_index in range(extra_samples):
        target_start = table_a_insert_at + extra_index * 12
        row_offset = target_start - 86
        for offset in range(12):
            _copy_row_all(sheet, 86 + offset, target_start + offset)
        _copy_merged_ranges_with_offset(sheet, 86, 97, row_offset)

    table_b_insert_at = 113 + table_a_rows
    table_b_rows = extra_samples * 2
    shifted = _shift_merged_ranges_for_insert(sheet, table_b_insert_at, table_b_rows)
    sheet.insert_rows(table_b_insert_at, table_b_rows)
    _restore_shifted_merged_ranges(sheet, shifted, table_b_rows)

    for extra_index in range(extra_samples):
        target_start = table_b_insert_at + extra_index * 2
        for offset in range(2):
            _copy_row_all(sheet, 111 + offset + table_a_rows, target_start + offset)


def _spd03012_table_a_start(sample_index: int) -> int:
    if sample_index <= 5:
        return SPD03012_TABLE_A_START_ROWS[sample_index]
    return 98 + (sample_index - 6) * 12


def _spd03012_table_b_start(sample_index: int, sample_count: int) -> int:
    if sample_count < 5:
        base = SPD03012_TABLE_B_START_ROW - (5 - sample_count) * 12
    else:
        base = SPD03012_TABLE_B_START_ROW + max(sample_count - 5, 0) * 12
    return base + (sample_index - 1) * 2


def _repair_spd03012_header(sheet) -> None:
    blue_fill = PatternFill("solid", fgColor="00338D")
    white_font = Font(name="Arial", bold=True, color=WHITE, size=10)
    subtitle_font = Font(name="Arial", bold=True, italic=True, color=WHITE, size=8)

    for row in (1, 2, 3):
        for col in range(1, 20):
            cell = sheet.cell(row, col)
            cell.fill = blue_fill
            cell.border = Border()
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    for merged_range in list(sheet.merged_cells.ranges):
        if merged_range.min_row <= 3 and merged_range.max_row >= 1:
            sheet.unmerge_cells(str(merged_range))

    for range_ref in ("B1:S1", "B2:S2", "B3:S3"):
        sheet.merge_cells(range_ref)

    sheet["B2"] = "kpmg    存货的固定加工成本 (标准成本法)  - 重新计算"
    sheet["B3"] = "（本实质性程序模板用于测试标准成本法下的存货固定加工成本。）"
    sheet["B2"].font = white_font
    sheet["B3"].font = subtitle_font
    sheet["B2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    sheet["B3"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 8
    sheet.row_dimensions[2].height = 24
    sheet.row_dimensions[3].height = 20


def _populate_template_spd03012(sheet, source_folder: Path, samples: list[Spd03015Sample]) -> None:
    _repair_spd03012_header(sheet)
    _ensure_spd03012_sample_rows(sheet, len(samples))
    data = _build_spd03012_data(source_folder, samples)
    for sample_index, sample_data in enumerate(data, 1):
        block_start = _spd03012_table_a_start(sample_index)
        sample_no = sample_data["sample"]
        screenshot_type = _document_type_for_report(source_folder, sample_no, "CO03", prefer_image=True)
        co03_table_type = _document_type_for_report(source_folder, sample_no, "CO03", prefer_image=False)

        for expense_index, expense_data in enumerate(sample_data["expenses"]):
            row = block_start + expense_index * 6
            sheet.cell(row, 2, f"样本{sample_data['sample']}")
            sheet.cell(row, 3, sample_data["order"])
            sheet.cell(row, 4, sample_data["cost_center"])
            sheet.cell(row, 5, expense_data["expense"])
            sheet.cell(row, 6, expense_data["plan"])
            sheet.cell(row, 7, expense_data["actual"])
            sheet.cell(row, 8, expense_data["variance"])
            sheet.cell(row, 9, expense_data["actual_rate"])
            sheet.cell(row, 10, f"=L{row}*N{row}-F{row}")
            sheet.cell(row, 11, f"=H{row}-J{row}")
            sheet.cell(row, 12, expense_data["qty"])
            sheet.cell(row, 13, "不适用")
            sheet.cell(row, 14, expense_data["actual_absorption_rate"])
            sheet.cell(row, 15, f"=N{row}-I{row}")
            sheet.cell(row, 18, f"=L{row}")
            for doc_offset in range(6):
                sheet.cell(row + doc_offset, 17, co03_table_type if doc_offset == 4 else screenshot_type)

        table_b_start = _spd03012_table_b_start(sample_index, len(data))
        for expense_index, expense_data in enumerate(sample_data["expenses"]):
            table_a_row = block_start + expense_index * 6
            row = table_b_start + expense_index
            qty = expense_data["qty"]
            standard_unit_cost = expense_data["plan"] / qty if qty else expense_data["plan_rate"]
            standard_total = standard_unit_cost * qty if qty else expense_data["plan"]

            sheet.cell(row, 2, f"样本{sample_data['sample']}")
            sheet.cell(row, 3, sample_data["product_id"])
            sheet.cell(row, 4, sample_data["cost_center"])
            sheet.cell(row, 5, sample_data["order"])
            sheet.cell(row, 6, expense_data["expense"])
            sheet.cell(row, 7, standard_unit_cost)
            sheet.cell(row, 8, qty)
            sheet.cell(row, 9, standard_total)
            sheet.cell(row, 10, expense_data["actual"])
            sheet.cell(row, 11, f"=J{row}-I{row}")
            sheet.cell(row, 12, f"=H{table_a_row}")
            sheet.cell(row, 13, f"=L{row}-K{row}")


def _normalize_programs(programs: str | list[str] | tuple[str, ...] | set[str] | None) -> set[str]:
    if programs is None:
        return set(PROGRAM_SHEETS)
    if isinstance(programs, str):
        values = [programs]
    else:
        values = list(programs)
    selected = {str(value).upper().replace("_IRM(SAP)", "").strip() for value in values}
    unsupported = selected - set(PROGRAM_SHEETS)
    if unsupported:
        raise ValueError(f"Unsupported SPD program(s): {', '.join(sorted(unsupported))}")
    return selected


def _build_from_template(
    samples: list[Spd03015Sample],
    source_folder: Path,
    programs: str | list[str] | tuple[str, ...] | set[str] | None = None,
) -> Workbook:
    selected = _normalize_programs(programs)
    workbook = load_workbook(TEMPLATE_PATH, data_only=False)
    if "SPD03012" in selected and "SPD03012-IRM(SAP)" in workbook.sheetnames:
        _populate_template_spd03012(workbook["SPD03012-IRM(SAP)"], source_folder, samples)
    if "SPD03014" in selected and "SPD03014_IRM(SAP)" in workbook.sheetnames:
        _populate_template_spd03014(workbook["SPD03014_IRM(SAP)"], source_folder, samples)
    if "SPD03015" in selected and "SPD03015_IRM(SAP)" in workbook.sheetnames:
        _populate_template_spd(workbook["SPD03015_IRM(SAP)"], samples)

    keep_sheets = {PROGRAM_SHEETS[program] for program in selected}
    for sheet_name in list(workbook.sheetnames):
        if sheet_name == "CKM3" or sheet_name in PROGRAM_SHEETS.values() and sheet_name not in keep_sheets:
            del workbook[sheet_name]
    if "CKM3" in workbook.sheetnames:
        del workbook["CKM3"]
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    return workbook


def build_spd03015_workbook(
    source_folder: Path,
    program: str | list[str] | tuple[str, ...] | set[str] | None = None,
) -> Workbook:
    samples = _discover_samples(source_folder)
    if TEMPLATE_PATH.exists():
        return _build_from_template(samples, source_folder, program)

    workbook = Workbook()
    ck_rows = _build_ckm3_sheet(workbook, samples)
    _build_spd_sheet(workbook, samples, ck_rows)
    if "CKM3" in workbook.sheetnames:
        del workbook["CKM3"]
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    return workbook


def build_spd03015_bytes(
    source_folder: Path,
    program: str | list[str] | tuple[str, ...] | set[str] | None = None,
) -> bytes:
    workbook = build_spd03015_workbook(source_folder, program)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()
