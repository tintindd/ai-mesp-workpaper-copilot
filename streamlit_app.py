from __future__ import annotations

import shutil
import sys
import tempfile
import zipfile
from io import BytesIO
from pathlib import Path
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
import streamlit as st


PROJECT_ROOT = Path(__file__).resolve().parent
SCRIPTS_ROOT = PROJECT_ROOT / "scripts"
if str(SCRIPTS_ROOT) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_ROOT))

from mesp_automation_engine import REQUIRED_FIELDS, analyze_folder, analyze_workbook  # noqa: E402
from deepseek_client import (  # noqa: E402
    clean_ocr_text_with_deepseek,
    load_deepseek_config,
    normalize_filename_with_deepseek,
    test_deepseek_connection,
)
import ocr_client  # noqa: E402
from ckm3_table_builder import build_ckm3_workbook_bytes, extract_ckm3_rows  # noqa: E402
from spd03015_exporter import build_spd03015_bytes  # noqa: E402
from supporting_exporter import build_supporting_bytes  # noqa: E402


is_image_file = ocr_client.is_image_file
load_online_ocr_config = ocr_client.load_online_ocr_config
online_ocr_available = ocr_client.online_ocr_available
recognize_uploaded_image = ocr_client.recognize_uploaded_image
recognize_co03_order_id = getattr(ocr_client, "recognize_co03_order_id", None)
recognize_co03_product_id = getattr(ocr_client, "recognize_co03_product_id", None)
recognize_ckm3_material_id = getattr(ocr_client, "recognize_ckm3_material_id", None)
recognize_ksbt_cost_center = getattr(ocr_client, "recognize_ksbt_cost_center", None)


st.set_page_config(
    page_title="AI-MESP Workpaper Copilot",
    page_icon="📊",
    layout="wide",
)

query_theme = st.query_params.get("theme", st.session_state.get("ui_theme", "light"))
if isinstance(query_theme, list):
    query_theme = query_theme[0] if query_theme else "light"
if query_theme not in {"light", "dark"}:
    query_theme = "light"

st.session_state["ui_theme"] = query_theme
theme_is_dark = query_theme == "dark"
theme_class = "theme-dark" if theme_is_dark else "theme-light"
next_theme = "light" if theme_is_dark else "dark"
theme_label = "主题切换"

st.markdown(
    """
    <style>
    [data-testid="stHeader"] { display: none; }
    [data-testid="stSidebar"] { display: none; }
    .block-container {
        max-width: none;
        padding: 0 0 3rem 0;
    }
    .stApp {
        background: #f4f7fb;
        color: #172033;
    }
    .mesp-hero {
        padding: 2.3rem 4.8rem 2.1rem;
        background: linear-gradient(120deg, #00338d 0%, #005eb8 65%, #00a3e0 100%);
        color: #fff;
        display: flex;
        align-items: center;
        justify-content: space-between;
        gap: 2rem;
        border-bottom: 1px solid rgba(255,255,255,0.14);
    }
    .mesp-hero h1 {
        margin: 0 0 .45rem;
        font-size: 2.1rem;
        line-height: 1.1;
        font-weight: 800;
        letter-spacing: 0;
    }
    .mesp-hero p {
        margin: 0;
        font-size: 1.02rem;
        opacity: .9;
    }
    .mesp-badge {
        border: 1px solid rgba(255,255,255,.38);
        background: rgba(255,255,255,.12);
        border-radius: .55rem;
        padding: .7rem 1rem;
        font-weight: 700;
        white-space: nowrap;
    }
    .mesp-shell {
        width: min(92rem, calc(100vw - 4.5rem));
        margin: 2rem auto 0;
    }
    .step-card {
        background: #fff;
        border: 1px solid #d7e1ee;
        border-radius: .65rem;
        padding: 1.35rem;
        box-shadow: 0 1rem 2rem rgba(20,32,51,.07);
    }
    .step-item {
        display: grid;
        grid-template-columns: 2.2rem 1fr;
        gap: .9rem;
        align-items: start;
        padding: 1rem .75rem;
        border-radius: .5rem;
        margin-bottom: .45rem;
    }
    .step-item.active { background: #edf4ff; color: #00338d; }
    .step-no {
        width: 2rem;
        height: 2rem;
        border-radius: 50%;
        border: 1px solid #6f819d;
        display: grid;
        place-items: center;
        font-weight: 800;
        color: #00338d;
    }
    .step-item strong { display: block; font-size: 1rem; margin-bottom: .25rem; }
    .step-item span { color: #5c6f8c; font-size: .9rem; }
    .main-title h2 {
        margin: .2rem 0 .45rem;
        font-size: 1.65rem;
        font-weight: 850;
        letter-spacing: 0;
    }
    .main-title p {
        color: #587091;
        margin: 0 0 1.3rem;
        line-height: 1.6;
    }
    div[data-testid="stVerticalBlockBorderWrapper"] {
        border-color: #d7e1ee;
        border-radius: .65rem;
        box-shadow: 0 1rem 2rem rgba(20,32,51,.06);
        background: #fff;
    }
    div[data-testid="stMetric"] {
        background: #fff;
        border: 1px solid #d7e1ee;
        border-radius: .55rem;
        padding: 1rem 1.1rem;
        box-shadow: 0 .6rem 1.4rem rgba(20,32,51,.045);
    }
    div[data-testid="stMetric"] label { color: #587091; }
    div[data-testid="stMetricValue"] { color: #00338d; }
    .program-help {
        display: grid;
        grid-template-columns: repeat(3, minmax(0, 1fr));
        gap: .85rem;
        margin: .2rem 0 1rem;
    }
    .program-card {
        border: 1px solid #d7e1ee;
        border-radius: .55rem;
        padding: .9rem 1rem;
        min-height: 6.4rem;
        background: #fff;
    }
    .program-card strong { display: block; margin-bottom: .35rem; }
    .program-code { color: #00338d; font-weight: 850; }
    .pill {
        display: inline-block;
        padding: .22rem .55rem;
        border-radius: .4rem;
        border: 1px solid #cfd9e7;
        background: #eef3fa;
        color: #172033;
        margin: .45rem .25rem 0 0;
        font-size: .82rem;
    }
    .upload-rules {
        border: 1px solid #cfe1f7;
        background: #eef6ff;
        color: #17446f;
        border-radius: .55rem;
        padding: 1rem 1.15rem;
        line-height: 1.7;
        margin: .6rem 0 1rem;
    }
    .upload-rules code {
        background: rgba(255,255,255,.78);
        border: 1px solid #cfd9e7;
        border-radius: .35rem;
        padding: .08rem .32rem;
        color: #00338d;
    }
    .stButton > button, .stDownloadButton > button {
        border-radius: .5rem;
        font-weight: 800;
        min-height: 2.7rem;
    }
    .stButton > button[kind="primary"],
    .stDownloadButton > button[kind="primary"] {
        background: #e7f1ff !important;
        border-color: #9fc5f8 !important;
        color: #00338d !important;
    }
    .stButton > button[kind="primary"]:hover,
    .stDownloadButton > button[kind="primary"]:hover {
        background: #d8eaff !important;
        border-color: #6ea8ee !important;
        color: #002f6c !important;
    }
    .stButton > button[kind="primary"]:focus,
    .stDownloadButton > button[kind="primary"]:focus {
        box-shadow: 0 0 0 .18rem rgba(0, 94, 184, .18) !important;
    }
    [data-testid="stRadio"] [role="radiogroup"] label span:first-child {
        border-color: #7fb3f0 !important;
    }
    [data-testid="stRadio"] [role="radiogroup"] label span:first-child:has(input:checked) {
        background-color: #005eb8 !important;
        border-color: #005eb8 !important;
    }
    .fork-button {
        position: fixed;
        top: .45rem;
        right: .55rem;
        z-index: 9999;
        display: inline-flex;
        align-items: center;
        gap: .42rem;
        padding: .45rem .72rem;
        border: 1px solid rgba(255,255,255,.25);
        border-radius: .2rem;
        background: rgba(8, 12, 20, .72);
        color: #fff !important;
        font-size: .78rem;
        font-weight: 800;
        text-decoration: none !important;
        line-height: 1;
        box-shadow: 0 .25rem .75rem rgba(0,0,0,.22);
        backdrop-filter: blur(6px);
    }
    .fork-button:hover {
        background: rgba(0, 51, 141, .92);
        border-color: rgba(255,255,255,.45);
    }
    .fork-button svg {
        width: .95rem;
        height: .95rem;
        fill: currentColor;
    }
    .theme-button {
        position: fixed;
        top: .45rem;
        right: 5.05rem;
        z-index: 10000;
        display: inline-flex;
        align-items: center;
        justify-content: center;
        padding: .45rem .72rem;
        border: 1px solid rgba(255,255,255,.25);
        border-radius: .2rem;
        background: rgba(8, 12, 20, .72);
        color: #fff !important;
        font-size: .78rem;
        font-weight: 800;
        text-decoration: none !important;
        line-height: 1;
        box-shadow: 0 .25rem .75rem rgba(0,0,0,.22);
        backdrop-filter: blur(6px);
    }
    .theme-button:hover {
        background: rgba(0, 51, 141, .92);
        border-color: rgba(255,255,255,.45);
    }
    .theme-dark {
        background: #05070b;
        color: #f8fafc;
        min-height: 100vh;
    }
    .theme-dark .mesp-hero {
        background: linear-gradient(120deg, #030712 0%, #06142f 54%, #00338d 100%);
        border-bottom-color: rgba(255,255,255,.08);
    }
    .theme-dark .step-card,
    .theme-dark div[data-testid="stVerticalBlockBorderWrapper"],
    .theme-dark .program-card,
    .theme-dark div[data-testid="stMetric"] {
        background: #0b0f19 !important;
        border-color: #20293a !important;
        box-shadow: 0 1rem 2rem rgba(0,0,0,.38) !important;
    }
    .theme-dark .step-item.active {
        background: #0e2444;
        color: #93c5fd;
    }
    .theme-dark .step-no {
        border-color: #4b638a;
        color: #93c5fd;
    }
    .theme-dark .step-item span,
    .theme-dark .main-title p,
    .theme-dark div[data-testid="stMetric"] label {
        color: #9aa8bd !important;
    }
    .theme-dark .program-code,
    .theme-dark div[data-testid="stMetricValue"] {
        color: #60a5fa !important;
    }
    .theme-dark .pill {
        background: #111827;
        border-color: #2a3548;
        color: #dbeafe;
    }
    .theme-dark .upload-rules {
        background: #0b1728;
        border-color: #1e3a5f;
        color: #bfdbfe;
    }
    .theme-dark .upload-rules code {
        background: #0f172a;
        border-color: #2a3548;
        color: #93c5fd;
    }
    .theme-dark input,
    .theme-dark textarea,
    .theme-dark select,
    .theme-dark [data-baseweb="select"] > div,
    .theme-dark [data-testid="stFileUploaderDropzone"] {
        background: #090d15 !important;
        color: #f8fafc !important;
        border-color: #263246 !important;
    }
    .theme-dark [data-testid="stDataFrame"],
    .theme-dark [data-testid="stTable"] {
        background: #0b0f19 !important;
    }
    @media (max-width: 900px) {
        .mesp-hero { padding: 1.6rem 1.4rem; align-items: flex-start; flex-direction: column; }
        .mesp-shell { width: calc(100vw - 1.5rem); margin-top: 1rem; }
        .program-help { grid-template-columns: 1fr; }
        .fork-button { top: .35rem; right: .35rem; }
        .theme-button { top: .35rem; right: 4.85rem; }
    }
    </style>
    """,
    unsafe_allow_html=True,
)

if theme_is_dark:
    st.markdown(
        """
        <style>
        .stApp,
        [data-testid="stAppViewContainer"] {
            background: #05070b !important;
            color: #f8fafc !important;
        }
        [data-testid="stHeader"] {
            background: transparent !important;
        }
        .block-container {
            background: #05070b !important;
        }
        .mesp-hero {
            background: linear-gradient(120deg, #020617 0%, #09204d 56%, #005eb8 100%) !important;
            border-bottom: 1px solid rgba(255,255,255,.08) !important;
        }
        .mesp-shell {
            color: #f8fafc !important;
        }
        .step-card,
        div[data-testid="stVerticalBlockBorderWrapper"],
        .program-card,
        div[data-testid="stMetric"] {
            background: #0f172a !important;
            border-color: #243047 !important;
            box-shadow: 0 1rem 2rem rgba(0,0,0,.34) !important;
        }
        .step-item.active {
            background: #102647 !important;
            color: #dbeafe !important;
        }
        .step-no {
            border-color: #5b7aaa !important;
            color: #93c5fd !important;
            background: rgba(96,165,250,.08) !important;
        }
        .step-item strong,
        .main-title h2,
        .program-card strong,
        h1, h2, h3, h4, h5, h6,
        label,
        [data-testid="stMarkdownContainer"] {
            color: #f8fafc !important;
        }
        .step-item span,
        .main-title p,
        div[data-testid="stMetric"] label,
        [data-testid="stWidgetLabel"] p,
        [data-testid="stFileUploader"] small {
            color: #b6c4d8 !important;
        }
        [data-testid="stFileUploader"] small,
        [data-testid="stFileUploaderDropzone"] small,
        [data-testid="stFileUploaderDropzone"] span {
            color: #dbeafe !important;
            opacity: 1 !important;
        }
        .program-code,
        div[data-testid="stMetricValue"] {
            color: #60a5fa !important;
        }
        .pill {
            background: #172033 !important;
            border-color: #334155 !important;
            color: #dbeafe !important;
        }
        .upload-rules {
            background: #0b1f37 !important;
            border-color: #1d4f82 !important;
            color: #d8ecff !important;
        }
        .upload-rules strong {
            color: #ffffff !important;
        }
        .upload-rules code {
            background: #07111f !important;
            border-color: #2d4567 !important;
            color: #93c5fd !important;
        }
        input,
        textarea,
        select,
        [data-baseweb="select"] > div,
        [data-testid="stFileUploaderDropzone"] {
            background: #111827 !important;
            color: #f8fafc !important;
            border-color: #334155 !important;
        }
        [data-baseweb="select"] span,
        [data-baseweb="select"] div,
        [data-testid="stTextInput"] input {
            color: #f8fafc !important;
        }
        [data-testid="stFileUploaderDropzone"] button,
        .stButton > button,
        .stDownloadButton > button {
            background: #172033 !important;
            border-color: #334155 !important;
            color: #f8fafc !important;
        }
        .stButton > button[kind="primary"],
        .stDownloadButton > button[kind="primary"] {
            background: #005eb8 !important;
            border-color: #2b8edb !important;
            color: #ffffff !important;
        }
        [data-testid="stRadio"] label,
        [data-testid="stRadio"] p {
            color: #dbeafe !important;
        }
        [data-testid="stDataFrame"],
        [data-testid="stTable"] {
            background: #0f172a !important;
            color: #f8fafc !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def save_uploaded_files(uploaded_files, target_dir: Path) -> int:
    saved = 0
    for uploaded_file in uploaded_files:
        name = uploaded_file.name
        if name.lower().endswith(".zip"):
            zip_path = target_dir / name
            zip_path.write_bytes(uploaded_file.getbuffer())
            with zipfile.ZipFile(zip_path) as archive:
                for member in archive.infolist():
                    if member.is_dir():
                        continue
                    member_path = Path(member.filename)
                    if member_path.is_absolute() or ".." in member_path.parts:
                        continue
                    destination = target_dir / member_path
                    destination.parent.mkdir(parents=True, exist_ok=True)
                    with archive.open(member) as source, destination.open("wb") as output:
                        shutil.copyfileobj(source, output)
                    saved += 1
            zip_path.unlink(missing_ok=True)
            continue

        destination = target_dir / Path(name).name
        destination.write_bytes(uploaded_file.getbuffer())
        saved += 1
    return saved


def extract_zip_bytes(zip_bytes: bytes, target_dir: Path) -> int:
    saved = 0
    with zipfile.ZipFile(BytesIO(zip_bytes)) as archive:
        for member in archive.infolist():
            if member.is_dir():
                continue
            member_path = Path(member.filename)
            if member_path.is_absolute() or ".." in member_path.parts:
                continue
            destination = target_dir / member_path
            destination.parent.mkdir(parents=True, exist_ok=True)
            with archive.open(member) as source, destination.open("wb") as output:
                shutil.copyfileobj(source, output)
            saved += 1
    return saved


def build_analysis_bundle_from_folder(source_dir: Path, period: str, program: str) -> dict:
    result = analyze_folder(source_dir, period=period, program=program)
    supporting_bytes = build_supporting_bytes(result, source_dir)
    spp_dir = source_dir / "_generated_spp"
    spp_dir.mkdir(parents=True, exist_ok=True)
    (spp_dir / "AI-MESP_SPP_Supporting.xlsx").write_bytes(supporting_bytes)
    summary = result.get("summary", {})
    selected_spd_bytes = (
        build_spd03015_bytes(source_dir, program=program, period=period)
        if summary.get("sample_count", 0) > 0 and summary.get("recognized_file_count", 0) > 0
        else b""
    )
    return {
        "result": result,
        "program": program,
        "supporting_bytes": supporting_bytes,
        "selected_spd_bytes": selected_spd_bytes,
    }


def run_analysis_from_cleaned_zip(zip_bytes: bytes | None, period: str, program: str) -> None:
    if not zip_bytes:
        st.session_state["filename_cleanup_error"] = "请先运行文件名清洗并生成标准命名文件包。"
        return
    try:
        with tempfile.TemporaryDirectory(prefix="mesp_cleaned_zip_") as temp:
            temp_dir = Path(temp)
            saved_count = extract_zip_bytes(zip_bytes, temp_dir)
            if saved_count == 0:
                st.session_state["filename_cleanup_error"] = "标准命名文件包为空，无法计算。"
                return
            st.session_state["analysis_bundle"] = build_analysis_bundle_from_folder(temp_dir, period, program)
            st.session_state["current_step"] = 4
    except Exception as exc:
        st.session_state["filename_cleanup_error"] = str(exc)


def render_issues(issues: list[dict]) -> None:
    if not issues:
        st.success("未发现异常或待补充事项。")
        return
    st.dataframe(
        [
            {
                "类型": item.get("type"),
                "样本": item.get("sample"),
                "问题": item.get("problem"),
                "建议": item.get("suggestion"),
                "状态": item.get("status"),
            }
            for item in issues
        ],
        use_container_width=True,
        hide_index=True,
    )


def render_workbooks(items: list[dict]) -> None:
    if not items:
        st.info("暂无 SAP 表格映射结果。")
        return
    st.dataframe(
        [
            {
                "样本": item.get("sample"),
                "报表": item.get("report"),
                "订单号": item.get("order") or "-",
                "文件": item.get("file"),
                "Sheet": item.get("sheet"),
                "行数": item.get("row_count"),
                "映射字段数": len(item.get("mapping") or {}),
                "缺失字段数": len(item.get("missing_fields") or []),
            }
            for item in items
        ],
        use_container_width=True,
        hide_index=True,
    )


def render_trace(items: list[dict]) -> None:
    if not items:
        st.info("暂无 Evidence Traceability 结果。")
        return
    st.dataframe(
        [
            {
                "样本": item.get("sample"),
                "报表": item.get("report"),
                "底稿字段": item.get("workpaper_field"),
                "来源文件": item.get("source_file"),
                "来源 Sheet": item.get("source_sheet"),
                "来源字段": item.get("source_header"),
                "来源列": item.get("source_column"),
            }
            for item in items[:300]
        ],
        use_container_width=True,
        hide_index=True,
    )


def mark_scenario_started() -> None:
    st.session_state["scenario_started"] = True
    st.session_state["params_confirmed"] = False
    st.session_state["ckm3_ocr_confirmed"] = False
    st.session_state["filename_cleanup_confirmed"] = False
    st.session_state["has_uploaded_files"] = False
    st.session_state.pop("analysis_bundle", None)


def go_to_step(step: int) -> None:
    if step <= get_unlocked_step():
        st.session_state["current_step"] = step


def go_next_from_scenario() -> None:
    st.session_state["scenario_started"] = True
    st.session_state["current_step"] = 2


PROGRAM_OPTIONS = {
    "SPD03012 - 标准成本法固定加工成本": "SPD03012",
    "SPD03014 - 标准成本法可变加工成本": "SPD03014",
    "SPD03015 - 存货成本差异分摊": "SPD03015",
}

PROGRAM_REQUIRED_REPORTS = {
    "SPD03012": ["CO03", "KSBT", "3611"],
    "SPD03014": ["CO03", "KSBT", "3611"],
    "SPD03015": ["CKM3", "3611"],
}

PROGRAM_REQUIRES_ORDER_ID = {
    "SPD03012": True,
    "SPD03014": True,
    "SPD03015": False,
}


def required_reports_for_program(program: str | None = None) -> list[str]:
    program_code = str(program or st.session_state.get("test_program") or "SPD03012").upper()
    return PROGRAM_REQUIRED_REPORTS.get(program_code, ["CO03", "KSBT", "3611"])


def program_requires_order_id(program: str | None = None) -> bool:
    program_code = str(program or st.session_state.get("test_program") or "SPD03012").upper()
    return PROGRAM_REQUIRES_ORDER_ID.get(program_code, True)


def sync_test_program_from_label() -> None:
    st.session_state["test_program"] = PROGRAM_OPTIONS.get(
        st.session_state.get("test_program_label"),
        "SPD03012",
    )
    mark_scenario_started()


def go_next_from_params() -> None:
    st.session_state["scenario_started"] = True
    st.session_state["params_confirmed"] = True
    st.session_state["current_step"] = 3


def get_unlocked_step() -> int:
    if st.session_state.get("analysis_bundle"):
        return 4
    if st.session_state.get("params_confirmed"):
        return 3
    if st.session_state.get("scenario_started"):
        return 2
    return 1


def get_active_step() -> int:
    if "current_step" not in st.session_state:
        st.session_state["current_step"] = 1
    unlocked_step = get_unlocked_step()
    if st.session_state["current_step"] > unlocked_step:
        st.session_state["current_step"] = unlocked_step
    return st.session_state["current_step"]


def step_class(step: int, active_step: int) -> str:
    return "step-item active" if step == active_step else "step-item"


def run_deepseek_connection_test() -> None:
    if not deepseek_config:
        st.session_state["deepseek_status"] = {
            "ok": False,
            "message": "DeepSeek API Key 未配置。",
        }
        return
    try:
        response = test_deepseek_connection(deepseek_config)
        st.session_state["deepseek_status"] = {
            "ok": True,
            "message": f"连接成功：{response.strip()}",
        }
    except Exception as exc:
        st.session_state["deepseek_status"] = {
            "ok": False,
            "message": str(exc),
        }


def display_missing_fields(report: str, missing_fields: list[str]) -> list[str]:
    aliases = REQUIRED_FIELDS.get(report, {})
    return [(aliases.get(field) or [field])[0] for field in missing_fields]


def run_filename_cleanup(uploaded_files) -> None:
    if not deepseek_config:
        st.session_state["filename_cleanup_results"] = []
        st.session_state["filename_cleanup_error"] = "DeepSeek API Key 未配置。"
        return

    results = []
    try:
        for uploaded_file in uploaded_files or []:
            results.append(normalize_filename_with_deepseek(uploaded_file.name, deepseek_config))
        st.session_state["filename_cleanup_results"] = results
        st.session_state.pop("filename_cleanup_error", None)
    except Exception as exc:
        st.session_state["filename_cleanup_error"] = str(exc)


def _cleanup_evidence_kind(filename: str) -> str:
    extension = Path(filename).suffix.lower().lstrip(".")
    if extension in {"xlsx", "xlsm", "xls", "csv"}:
        return "表格"
    if extension in {"png", "jpg", "jpeg", "pdf"}:
        return "截图"
    return "未知"


def _standard_cleanup_filename(
    *,
    sample_no: str,
    order_id: str,
    material_id: str,
    product_id: str = "",
    cost_center: str = "",
    report: str,
    evidence_kind: str,
    extension: str,
    include_order_id: bool = True,
) -> str:
    sample_text = str(sample_no or "1").strip()
    order_text = str(order_id or "待补充").strip()
    extension_text = extension.lstrip(".").lower()
    prefix = f"样本{sample_text}/{sample_text}."
    order_part = f"订单编号{order_text}-" if include_order_id else ""
    if report == "CKM3":
        material_text = str(material_id or "待补充").strip()
        return (
            f"{prefix}{order_part}"
            f"物料ID-{material_text}-CKM3-{evidence_kind}.{extension_text}"
        )
    if report == "CO03" and product_id:
        product_text = str(product_id).strip()
        return (
            f"{prefix}{order_part}"
            f"物料编码-{product_text}-CO03-{evidence_kind}.{extension_text}"
        )
    if report == "KSBT" and cost_center:
        cost_center_text = str(cost_center).strip()
        return (
            f"{prefix}{order_part}"
            f"成本中心-{cost_center_text}-KSBT-{evidence_kind}.{extension_text}"
        )
    return f"{prefix}{order_part}{report}-{evidence_kind}.{extension_text}"


def _deduplicate_zip_name(name: str, used_names: set[str]) -> str:
    if name not in used_names:
        used_names.add(name)
        return name
    path = Path(name)
    folder = path.parent.as_posix()
    stem = path.stem
    suffix = path.suffix
    counter = 2
    while True:
        candidate_name = f"{stem}-{counter}{suffix}"
        candidate = f"{folder}/{candidate_name}" if folder != "." else candidate_name
        if candidate not in used_names:
            used_names.add(candidate)
            return candidate
        counter += 1


def build_standard_named_zip_bytes(cleanup_results: list[dict]) -> bytes:
    output = BytesIO()
    used_names: set[str] = set()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
        for item in cleanup_results:
            file_bytes = item.get("file_bytes")
            standard_filename = item.get("standard_filename")
            if not file_bytes or not standard_filename:
                continue
            zip_name = _deduplicate_zip_name(str(standard_filename), used_names)
            archive.writestr(zip_name, file_bytes)
    return output.getvalue()


PARAMETER_COLUMNS = ["sample_no", "order_id", "product_id", "material_id", "cost_center"]


def _blank_evidence_params(sample_no: str = "1") -> dict:
    return {
        "sample_no": str(sample_no or "1").strip(),
        "order_id": "",
        "product_id": "",
        "material_id": "",
        "cost_center": "",
    }


def _normalize_evidence_params(row: dict | None, sample_no: str = "1") -> dict:
    source = row or {}
    normalized = _blank_evidence_params(source.get("sample_no") or sample_no)
    for key in PARAMETER_COLUMNS:
        normalized[key] = str(source.get(key) or normalized.get(key) or "").strip()
    normalized["sample_no"] = normalized["sample_no"] or str(sample_no or "1").strip()
    return normalized


def evidence_params_table() -> list[dict]:
    rows = st.session_state.get("evidence_params_table")
    if rows is None:
        legacy = st.session_state.get("evidence_params")
        rows = [_normalize_evidence_params(legacy)] if legacy else [_blank_evidence_params("1")]
        st.session_state["evidence_params_table"] = rows
    cleaned_rows = []
    seen = set()
    for row in rows:
        normalized = _normalize_evidence_params(row)
        sample_no = normalized["sample_no"] or "1"
        if sample_no in seen:
            continue
        seen.add(sample_no)
        cleaned_rows.append(normalized)
    if not cleaned_rows:
        cleaned_rows = [_blank_evidence_params("1")]
    st.session_state["evidence_params_table"] = sorted(
        cleaned_rows,
        key=lambda item: int(item["sample_no"]) if str(item["sample_no"]).isdigit() else str(item["sample_no"]),
    )
    return st.session_state["evidence_params_table"]


def upsert_evidence_params(params: dict) -> None:
    normalized = _normalize_evidence_params(params)
    rows = [row for row in evidence_params_table() if str(row.get("sample_no")) != normalized["sample_no"]]
    rows.append(normalized)
    st.session_state["evidence_params_table"] = sorted(
        rows,
        key=lambda item: int(item["sample_no"]) if str(item["sample_no"]).isdigit() else str(item["sample_no"]),
    )
    st.session_state["evidence_params"] = normalized
    st.session_state["current_param_sample_no"] = normalized["sample_no"]


def current_evidence_params(sample_no: str | None = None) -> dict:
    selected_sample = str(
        sample_no
        or st.session_state.get("current_param_sample_no")
        or st.session_state.get("cleanup_sample_no_select")
        or "1"
    ).strip()
    for row in evidence_params_table():
        if str(row.get("sample_no") or "").strip() == selected_sample:
            return _normalize_evidence_params(row, selected_sample)
    return _blank_evidence_params(selected_sample)


def delete_evidence_samples(samples: set[str]) -> None:
    if not samples:
        return
    st.session_state["evidence_params_table"] = [
        row
        for row in evidence_params_table()
        if str(row.get("sample_no") or "").strip() not in samples
    ] or [_blank_evidence_params("1")]
    st.session_state["filename_cleanup_results"] = [
        item
        for item in (st.session_state.get("filename_cleanup_results") or [])
        if str(item.get("sample_no") or "").strip() not in samples
    ]
    st.session_state["standard_named_zip_bytes"] = build_standard_named_zip_bytes(
        st.session_state["filename_cleanup_results"]
    )
    if st.session_state.get("current_param_sample_no") in samples:
        st.session_state["current_param_sample_no"] = st.session_state["evidence_params_table"][0]["sample_no"]


def apply_evidence_params_editor(edited_rows: list[dict]) -> None:
    kept_rows = [
        _normalize_evidence_params(row)
        for row in (edited_rows or [])
        if any(str(row.get(key) or "").strip() for key in PARAMETER_COLUMNS)
    ]
    st.session_state["evidence_params_table"] = kept_rows or [_blank_evidence_params("1")]


def evidence_sample_count() -> int:
    numeric_samples = []
    for row in evidence_params_table():
        sample_no = str(row.get("sample_no") or "").strip()
        if sample_no.isdigit():
            numeric_samples.append(int(sample_no))
    for item in st.session_state.get("filename_cleanup_results") or []:
        sample_no = str(item.get("sample_no") or "").strip()
        if sample_no.isdigit():
            numeric_samples.append(int(sample_no))
    return max([1, *numeric_samples])


def run_parameter_ocr(sample_no: str, co03_file, ckm3_file, ksbt_file) -> None:
    params = {
        "sample_no": str(sample_no or "1").strip(),
        "order_id": "",
        "product_id": "",
        "material_id": "",
        "cost_center": "",
    }
    preview_items = []
    st.session_state.pop("parameter_ocr_error", None)
    if not online_ocr_available(online_ocr_config):
        st.session_state["parameter_ocr_error"] = "Qwen-OCR 服务未配置，无法识别参数。"
        return
    try:
        progress = st.progress(0, text="准备识别样本参数...")
        status = st.empty()
        steps = max(sum(1 for item in [co03_file, ckm3_file, ksbt_file] if item), 1)
        done = 0

        def tick(message: str) -> None:
            nonlocal done
            status.caption(message)
            progress.progress(min(done / steps, 1.0), text=message)

        if co03_file:
            data = co03_file.getvalue()
            tick(f"正在识别 CO03 订单编号和物料编码：{co03_file.name}")
            if recognize_co03_order_id:
                params["order_id"] = recognize_co03_order_id(co03_file.name, data, online_ocr_config)
            if recognize_co03_product_id:
                params["product_id"] = recognize_co03_product_id(co03_file.name, data, online_ocr_config)
            preview_items.append({"label": "CO03截图", "source_file": co03_file.name, "file_bytes": data})
            done += 1
            tick("CO03 参数识别完成")

        if ckm3_file:
            data = ckm3_file.getvalue()
            tick(f"正在识别 CKM3 物料ID：{ckm3_file.name}")
            if recognize_ckm3_material_id:
                params["material_id"] = recognize_ckm3_material_id(ckm3_file.name, data, online_ocr_config)
            preview_items.append({"label": "CKM3截图", "source_file": ckm3_file.name, "file_bytes": data})
            done += 1
            tick("CKM3 参数识别完成")

        if ksbt_file:
            data = ksbt_file.getvalue()
            tick(f"正在识别 KSBT 成本中心：{ksbt_file.name}")
            if recognize_ksbt_cost_center:
                params["cost_center"] = recognize_ksbt_cost_center(ksbt_file.name, data, online_ocr_config)
            preview_items.append({"label": "KSBT截图", "source_file": ksbt_file.name, "file_bytes": data})
            done += 1
            tick("KSBT 参数识别完成")

        progress.progress(1.0, text="参数 OCR 识别完成")
        status.caption("参数 OCR 识别完成")
        upsert_evidence_params(params)
        st.session_state["parameter_ocr_previews"] = preview_items
        st.session_state["parameter_ocr_confirmed"] = False
    except Exception as exc:
        st.session_state["parameter_ocr_error"] = str(exc)


def apply_manual_params(sample_no: str, order_id: str, product_id: str, material_id: str, cost_center: str) -> None:
    upsert_evidence_params({
        "sample_no": str(sample_no or "1").strip(),
        "order_id": str(order_id or "").strip(),
        "product_id": str(product_id or "").strip(),
        "material_id": str(material_id or "").strip(),
        "cost_center": str(cost_center or "").strip(),
    })
    st.session_state["parameter_ocr_confirmed"] = True


def build_parameter_standard_files(params: dict, preview_items: list[dict]) -> list[dict]:
    results = []
    sample_no = params.get("sample_no") or "1"
    requires_order_id = program_requires_order_id()
    order_id = params.get("order_id") or ("待补充" if requires_order_id else "")
    for item in preview_items:
        source_file = item.get("source_file") or ""
        label = item.get("label") or ""
        if "CO03" in label:
            report = "CO03"
            product_id = params.get("product_id") or ""
            material_id = ""
            cost_center = ""
        elif "CKM3" in label:
            report = "CKM3"
            product_id = ""
            material_id = params.get("material_id") or ""
            cost_center = ""
        elif "KSBT" in label:
            report = "KSBT"
            product_id = ""
            material_id = ""
            cost_center = params.get("cost_center") or ""
        else:
            continue
        extension = Path(source_file).suffix.lower().lstrip(".") or "png"
        standard_filename = _standard_cleanup_filename(
            sample_no=sample_no,
            order_id=order_id,
            product_id=product_id,
            material_id=material_id,
            cost_center=cost_center,
            report=report,
            evidence_kind="截图",
            extension=extension,
            include_order_id=requires_order_id,
        )
        results.append(
            {
                "上传位置": report,
                "source_file": source_file,
                "sample_no": sample_no,
                "order_id": order_id,
                "product_id": product_id,
                "material_id": material_id,
                "cost_center": cost_center,
                "report_type": report,
                "evidence_kind": "截图",
                "extension": extension,
                "standard_filename": standard_filename,
                "field_status": "参数OCR截图",
                "missing_field_labels": [],
                "file_bytes": item.get("file_bytes"),
            }
        )
    return results


def confirm_parameter_standard_files(download_choice: bool) -> None:
    params = current_evidence_params()
    preview_items = st.session_state.get("parameter_ocr_previews") or []
    parameter_files = build_parameter_standard_files(params, preview_items)
    st.session_state["parameter_standard_files"] = parameter_files
    st.session_state["parameter_ocr_confirmed"] = True
    if not download_choice:
        existing = st.session_state.get("filename_cleanup_results") or []
        st.session_state["filename_cleanup_results"] = [*parameter_files, *existing]
        st.session_state["standard_named_zip_bytes"] = build_standard_named_zip_bytes(
            st.session_state["filename_cleanup_results"]
        )


def sample_completeness_rows(sample_count: int, cleanup_results: list[dict]) -> list[dict]:
    by_sample: dict[str, set[str]] = {str(index): set() for index in range(1, sample_count + 1)}
    for item in cleanup_results:
        sample = str(item.get("sample_no") or "")
        report = str(item.get("report_type") or "")
        kind = str(item.get("evidence_kind") or "")
        if sample and report:
            by_sample.setdefault(sample, set()).add(f"{report}-{kind}")
    rows = []
    required_reports = ["CO03", "KSBT", "3611", "CKM3"]
    for sample, files in sorted(by_sample.items(), key=lambda pair: int(pair[0]) if pair[0].isdigit() else 999):
        missing = []
        for report in required_reports:
            if not any(value.startswith(report) for value in files):
                missing.append(report)
        rows.append(
            {
                "样本": f"样本{sample}",
                "已识别文件": "、".join(sorted(files)) or "-",
                "缺失文件": "、".join(missing) or "完整",
                "是否可计算": "是" if not missing else "否",
            }
        )
    return rows


def sample_completeness_rows_v2(
    sample_count: int,
    cleanup_results: list[dict],
    required_reports: list[str] | None = None,
) -> list[dict]:
    by_sample: dict[str, set[str]] = {str(index): set() for index in range(1, sample_count + 1)}
    for item in cleanup_results:
        sample = str(item.get("sample_no") or "").strip()
        report = str(item.get("report_type") or "").strip()
        kind = str(item.get("evidence_kind") or "").strip()
        if not sample or not report:
            continue
        if "表" in kind:
            normalized_kind = "表格"
        elif "截" in kind or "图" in kind:
            normalized_kind = "截图"
        else:
            normalized_kind = kind or "未知"
        by_sample.setdefault(sample, set()).add(f"{report}-{normalized_kind}")

    rows = []
    required_reports = required_reports or ["CO03", "KSBT", "3611", "CKM3"]
    required_kinds = ["表格", "截图"]
    for sample, files in sorted(by_sample.items(), key=lambda pair: int(pair[0]) if pair[0].isdigit() else 999):
        missing = []
        for report in required_reports:
            for kind in required_kinds:
                expected = f"{report}-{kind}"
                if expected not in files:
                    missing.append(f"{report}{kind}")
        rows.append(
            {
                "样本": f"样本{sample}",
                "已识别/清洗文件": "、".join(sorted(files)) or "-",
                "缺失文件": "、".join(missing) or "完整",
                "是否可计算": "是" if not missing else "否",
            }
        )
    return rows


def add_missing_support_file(sample_no: str, report: str, evidence_kind: str, uploaded_file) -> None:
    st.session_state.pop("missing_support_file_error", None)
    if not uploaded_file:
        st.session_state["missing_support_file_error"] = "请先选择要补充上传的文件。"
        return

    params = current_evidence_params(sample_no)
    order_id = str(params.get("order_id") or "").strip()
    product_id = str(params.get("product_id") or "").strip()
    material_id = str(params.get("material_id") or "").strip()
    cost_center = str(params.get("cost_center") or "").strip()
    file_bytes = uploaded_file.getvalue()
    extension = Path(uploaded_file.name).suffix.lower().lstrip(".")
    expected_kind = _cleanup_evidence_kind(uploaded_file.name)
    if expected_kind in {"表格", "截图"} and expected_kind != evidence_kind:
        st.session_state["missing_support_file_error"] = (
            f"上传文件类型与选择不一致：当前选择为{evidence_kind}，但文件更像是{expected_kind}。"
        )
        return
    standard_filename = _standard_cleanup_filename(
        sample_no=sample_no,
        order_id=order_id,
        material_id=material_id,
        product_id=product_id,
        cost_center=cost_center,
        report=report,
        evidence_kind=evidence_kind,
        extension=extension,
        include_order_id=program_requires_order_id(),
    )

    field_status = "非表格文件，未检查字段"
    missing_fields = []
    if extension in {"xlsx", "xlsm", "csv"}:
        with tempfile.TemporaryDirectory(prefix="mesp_missing_file_") as temp:
            temp_path = Path(temp) / Path(uploaded_file.name).name
            temp_path.write_bytes(file_bytes)
            workbook_result = analyze_workbook(temp_path, report)
            missing_fields = workbook_result.get("missing_fields") or []
            field_status = "完整" if not missing_fields else "缺失字段"

    new_item = {
        "上传位置": report,
        "source_file": uploaded_file.name,
        "sample_no": str(sample_no),
        "order_id": order_id or ("待补充" if program_requires_order_id() else ""),
        "product_id": product_id,
        "material_id": material_id,
        "cost_center": cost_center,
        "report_type": report,
        "evidence_kind": evidence_kind,
        "extension": extension,
        "standard_filename": standard_filename,
        "field_status": field_status,
        "missing_fields": missing_fields,
        "missing_field_labels": display_missing_fields(report, missing_fields),
        "file_bytes": file_bytes,
        "source": "计算结果补充上传",
    }
    existing = st.session_state.get("filename_cleanup_results") or []
    filtered = [
        item
        for item in existing
        if not (
            str(item.get("sample_no") or "") == str(sample_no)
            and str(item.get("report_type") or "") == report
            and str(item.get("evidence_kind") or "") == evidence_kind
        )
    ]
    st.session_state["filename_cleanup_results"] = [*filtered, new_item]
    st.session_state["standard_named_zip_bytes"] = build_standard_named_zip_bytes(
        st.session_state["filename_cleanup_results"]
    )


def run_filename_cleanup_by_bucket(
    sample_no: str,
    order_id: str,
    bucket_files: dict[str, list],
) -> None:
    st.session_state["filename_cleanup_results"] = []
    st.session_state.pop("standard_named_zip_bytes", None)
    st.session_state.pop("filename_cleanup_error", None)

    results = []
    try:
        params = current_evidence_params(sample_no)
        sample_no = str(sample_no or params.get("sample_no") or "1").strip()
        order_id = str(order_id or params.get("order_id") or "").strip()
        param_product_id = str(params.get("product_id") or "").strip()
        param_material_id = str(params.get("material_id") or "").strip()
        param_cost_center = str(params.get("cost_center") or "").strip()
        total_files = sum(len(files or []) for files in bucket_files.values())
        requires_order_id = program_requires_order_id()
        missing_param_messages = []
        if requires_order_id and total_files and not order_id:
            missing_param_messages.append("订单编号")
        if bucket_files.get("CO03") and not param_product_id and not any(is_image_file(item.name) for item in bucket_files.get("CO03") or []):
            missing_param_messages.append("CO03物料编码")
        if bucket_files.get("CKM3") and not param_material_id and not any(is_image_file(item.name) for item in bucket_files.get("CKM3") or []):
            missing_param_messages.append("CKM3物料ID")
        if bucket_files.get("KSBT") and not param_cost_center and not any(is_image_file(item.name) for item in bucket_files.get("KSBT") or []):
            missing_param_messages.append("KSBT成本中心")
        if missing_param_messages:
            st.session_state["filename_cleanup_error"] = (
                "文件名清洗缺少关键参数："
                + "、".join(missing_param_messages)
                + "。请先在“参数获取”中进行OCR识别，或使用手工输入补充后再运行文件名清洗。"
            )
            return
        total_steps = max(total_files + 3, 1)
        step_index = 0
        progress_bar = st.progress(0, text="准备文件名清洗任务...")
        progress_status = st.empty()

        def update_progress(message: str) -> None:
            nonlocal step_index
            step_index += 1
            progress_bar.progress(min(step_index / total_steps, 1.0), text=message)
            progress_status.caption(message)

        with tempfile.TemporaryDirectory(prefix="mesp_filename_check_") as temp:
            temp_dir = Path(temp)
            detected_co03_product_id = param_product_id
            detected_ckm3_material_id = param_material_id
            detected_ksbt_cost_center = param_cost_center
            if online_ocr_available(online_ocr_config):
                for uploaded_file in bucket_files.get("CO03") or []:
                    if is_image_file(uploaded_file.name):
                        if recognize_co03_product_id:
                            update_progress(f"正在识别 CO03 物料编码：{uploaded_file.name}")
                            detected_co03_product_id = recognize_co03_product_id(
                                uploaded_file.name,
                                uploaded_file.getvalue(),
                                online_ocr_config,
                            )
                        if detected_co03_product_id:
                            break
                for uploaded_file in bucket_files.get("CKM3") or []:
                    if is_image_file(uploaded_file.name):
                        if recognize_ckm3_material_id:
                            update_progress(f"正在识别 CKM3 物料ID：{uploaded_file.name}")
                            detected_ckm3_material_id = recognize_ckm3_material_id(
                                uploaded_file.name,
                                uploaded_file.getvalue(),
                                online_ocr_config,
                            )
                        if detected_ckm3_material_id:
                            break
                for uploaded_file in bucket_files.get("KSBT") or []:
                    if is_image_file(uploaded_file.name):
                        if recognize_ksbt_cost_center:
                            update_progress(f"正在识别 KSBT 成本中心：{uploaded_file.name}")
                            detected_ksbt_cost_center = recognize_ksbt_cost_center(
                                uploaded_file.name,
                                uploaded_file.getvalue(),
                                online_ocr_config,
                            )
                        if detected_ksbt_cost_center:
                            break

            for report, files in bucket_files.items():
                for uploaded_file in files or []:
                    update_progress(f"正在清洗 {report} 文件名：{uploaded_file.name}")
                    file_bytes = uploaded_file.getvalue()
                    if deepseek_config:
                        cleaned = normalize_filename_with_deepseek(uploaded_file.name, deepseek_config)
                    else:
                        cleaned = {"source_file": uploaded_file.name}

                    extension = Path(uploaded_file.name).suffix.lower().lstrip(".")
                    evidence_kind = _cleanup_evidence_kind(uploaded_file.name)
                    cleaned["上传位置"] = report
                    cleaned["sample_no"] = str(sample_no or cleaned.get("sample_no") or "1").strip()
                    cleaned["report_type"] = report
                    cleaned["order_id"] = (
                        order_id.strip()
                        if order_id.strip()
                        else cleaned.get("order_id") or ("待补充" if requires_order_id else "")
                    )
                    if report == "CKM3":
                        detected_material_id = str(cleaned.get("material_id") or detected_ckm3_material_id or "").strip()
                        if (
                            not detected_material_id
                            and recognize_ckm3_material_id
                            and is_image_file(uploaded_file.name)
                            and online_ocr_available(online_ocr_config)
                        ):
                            progress_status.caption(f"正在补充识别 CKM3 物料ID：{uploaded_file.name}")
                            detected_material_id = recognize_ckm3_material_id(
                                uploaded_file.name,
                                file_bytes,
                                online_ocr_config,
                            )
                        cleaned["material_id"] = detected_material_id or "待补充"
                    if report == "CO03":
                        detected_product_id = str(cleaned.get("product_id") or detected_co03_product_id or "").strip()
                        if (
                            not detected_product_id
                            and recognize_co03_product_id
                            and is_image_file(uploaded_file.name)
                            and online_ocr_available(online_ocr_config)
                        ):
                            progress_status.caption(f"正在补充识别 CO03 物料编码：{uploaded_file.name}")
                            detected_product_id = recognize_co03_product_id(
                                uploaded_file.name,
                                file_bytes,
                                online_ocr_config,
                            )
                        cleaned["product_id"] = detected_product_id
                    if report == "KSBT":
                        detected_cost_center = str(cleaned.get("cost_center") or detected_ksbt_cost_center or "").strip()
                        if (
                            not detected_cost_center
                            and recognize_ksbt_cost_center
                            and is_image_file(uploaded_file.name)
                            and online_ocr_available(online_ocr_config)
                        ):
                            progress_status.caption(f"正在补充识别 KSBT 成本中心：{uploaded_file.name}")
                            detected_cost_center = recognize_ksbt_cost_center(
                                uploaded_file.name,
                                file_bytes,
                                online_ocr_config,
                            )
                        cleaned["cost_center"] = detected_cost_center
                    cleaned["evidence_kind"] = evidence_kind
                    cleaned["extension"] = extension
                    cleaned["standard_filename"] = _standard_cleanup_filename(
                        sample_no=cleaned["sample_no"],
                        order_id=cleaned["order_id"],
                        material_id=cleaned.get("material_id") or "",
                        product_id=cleaned.get("product_id") or "",
                        cost_center=cleaned.get("cost_center") or "",
                        report=report,
                        evidence_kind=evidence_kind,
                        extension=extension,
                        include_order_id=requires_order_id,
                    )
                    cleaned["file_bytes"] = file_bytes

                    field_status = "非表格文件，未检查字段"
                    missing_fields = []
                    if uploaded_file.name.lower().endswith((".xlsx", ".xlsm", ".csv")):
                        safe_name = Path(uploaded_file.name).name
                        temp_path = temp_dir / safe_name
                        temp_path.write_bytes(file_bytes)
                        report_for_check = report if report in {"CO03", "KSBT", "3611", "CKM3"} else cleaned["report_type"]
                        workbook_result = analyze_workbook(temp_path, report_for_check)
                        missing_fields = workbook_result.get("missing_fields") or []
                        field_status = "完整" if not missing_fields else "缺失字段"

                    cleaned["field_status"] = field_status
                    cleaned["missing_fields"] = missing_fields
                    cleaned["missing_field_labels"] = display_missing_fields(cleaned["report_type"], missing_fields)
                    cleaned["source"] = "DeepSeek 文件名清洗 + Excel 字段检查"
                    results.append(cleaned)
        progress_bar.progress(1.0, text="文件名清洗完成")
        progress_status.caption("文件名清洗完成")
        st.session_state["filename_cleanup_results"] = results
        st.session_state["standard_named_zip_bytes"] = build_standard_named_zip_bytes(results)
    except Exception as exc:
        st.session_state["filename_cleanup_error"] = str(exc)


def run_ckm3_ocr_to_excel(uploaded_files) -> None:
    st.session_state["ckm3_ocr_excels"] = []
    st.session_state.pop("ckm3_ocr_error", None)
    if not online_ocr_available(online_ocr_config):
        st.session_state["ckm3_ocr_error"] = "Qwen-OCR 服务未配置，无法识别 CKM3 截图。"
        return

    image_files = [item for item in uploaded_files or [] if is_image_file(item.name)]
    if not image_files:
        st.session_state["ckm3_ocr_error"] = "请上传 CKM3 截图文件（PNG/JPG/JPEG/PDF）。"
        return

    results = []
    try:
        progress_bar = st.progress(0, text="准备 CKM3 OCR 识别...")
        progress_status = st.empty()
        total = max(len(image_files), 1)
        for index, uploaded_file in enumerate(image_files, start=1):
            progress_bar.progress((index - 1) / total, text=f"正在识别 CKM3 截图 {index}/{total}：{uploaded_file.name}")
            progress_status.caption(f"正在识别 CKM3 截图 {index}/{total}：{uploaded_file.name}")
            uploaded_file.seek(0)
            ocr_result = recognize_uploaded_image(uploaded_file, online_ocr_config)
            ocr_text = ocr_result.get("text") or ""
            rows = extract_ckm3_rows(ocr_text)
            if not rows:
                results.append(
                    {
                        "source_file": uploaded_file.name,
                        "status": "未识别到 CKM3 表格行",
                        "row_count": 0,
                        "ocr_text": ocr_text,
                    }
                )
                continue
            output_name = f"{Path(uploaded_file.name).stem}-CKM3-表格.xlsx"
            results.append(
                {
                    "source_file": uploaded_file.name,
                    "status": "已生成",
                    "file_name": output_name,
                    "row_count": len(rows),
                    "bytes": build_ckm3_workbook_bytes(ocr_text),
                    "ocr_text": ocr_text,
                }
            )
        progress_bar.progress(1.0, text="CKM3 OCR 识别完成")
        progress_status.caption("CKM3 OCR 识别完成")
        st.session_state["ckm3_ocr_excels"] = results
    except Exception as exc:
        st.session_state["ckm3_ocr_error"] = str(exc)


def run_ocr_cleanup(uploaded_files) -> None:
    st.session_state["ocr_cleanup_results"] = []
    st.session_state.pop("ocr_cleanup_error", None)
    if not online_ocr_available(online_ocr_config):
        st.session_state["ocr_cleanup_error"] = "Qwen-OCR 服务未配置，无法运行图片 OCR。"
        return
    if not deepseek_config:
        st.session_state["ocr_cleanup_error"] = "DeepSeek API Key 未配置，无法清洗 OCR 文本。"
        return

    image_files = [item for item in uploaded_files or [] if is_image_file(item.name)]
    if not image_files:
        st.session_state["ocr_cleanup_error"] = "未找到可 OCR 的图片文件。"
        return

    results = []
    try:
        progress_bar = st.progress(0, text="准备 OCR 文本清洗...")
        progress_status = st.empty()
        total = max(len(image_files), 1)
        for index, uploaded_file in enumerate(image_files, start=1):
            progress_bar.progress((index - 1) / total, text=f"正在 OCR 清洗 {index}/{total}：{uploaded_file.name}")
            progress_status.caption(f"正在 OCR 清洗 {index}/{total}：{uploaded_file.name}")
            ocr_result = recognize_uploaded_image(uploaded_file, online_ocr_config)
            cleaned = clean_ocr_text_with_deepseek(
                uploaded_file.name,
                ocr_result.get("text") or "",
                deepseek_config,
            )
            results.append(
                {
                    "source_file": uploaded_file.name,
                    "ocr_text": ocr_result.get("text") or "",
                    "ocr_line_count": ocr_result.get("line_count", 0),
                    "cleaned": cleaned,
                }
            )
        progress_bar.progress(1.0, text="OCR 文本清洗完成")
        progress_status.caption("OCR 文本清洗完成")
        st.session_state["ocr_cleanup_results"] = results
    except Exception as exc:
        st.session_state["ocr_cleanup_error"] = str(exc)


def run_intelligent_cleanup(uploaded_files) -> None:
    st.session_state.pop("intelligent_cleanup_error", None)
    st.session_state.pop("intelligent_cleanup_excel", None)
    st.session_state["generated_support_excels"] = []
    st.session_state["filename_cleanup_results"] = []
    st.session_state["ocr_cleanup_results"] = []

    if not deepseek_config:
        st.session_state["intelligent_cleanup_error"] = "DeepSeek API Key 未配置，无法执行智能清洗。"
        return

    filename_results = []
    ocr_results = []
    try:
        for uploaded_file in uploaded_files or []:
            filename_results.append(normalize_filename_with_deepseek(uploaded_file.name, deepseek_config))

        if online_ocr_available(online_ocr_config):
            image_files = [item for item in uploaded_files or [] if is_image_file(item.name)]
            for uploaded_file in image_files:
                ocr_result = recognize_uploaded_image(uploaded_file, online_ocr_config)
                cleaned = clean_ocr_text_with_deepseek(
                    uploaded_file.name,
                    ocr_result.get("text") or "",
                    deepseek_config,
                )
                ocr_results.append(
                    {
                        "source_file": uploaded_file.name,
                        "ocr_text": ocr_result.get("text") or "",
                        "ocr_line_count": ocr_result.get("line_count", 0),
                        "cleaned": cleaned,
                    }
                )
                if (cleaned.get("report_type") or "").upper() == "CKM3":
                    ckm3_rows = extract_ckm3_rows(ocr_result.get("text") or "")
                    if ckm3_rows:
                        standard_filename = cleaned.get("standard_filename") or uploaded_file.name
                        file_name = standard_filename.replace("截图", "表格")
                        if not file_name.lower().endswith(".xlsx"):
                            file_name = f"{Path(file_name).stem}.xlsx"
                        st.session_state["generated_support_excels"].append(
                            {
                                "source_file": uploaded_file.name,
                                "file_name": file_name,
                                "report_type": "CKM3",
                                "row_count": len(ckm3_rows),
                                "bytes": build_ckm3_workbook_bytes(ocr_result.get("text") or ""),
                            }
                        )

        st.session_state["filename_cleanup_results"] = filename_results
        st.session_state["ocr_cleanup_results"] = ocr_results
        st.session_state["intelligent_cleanup_excel"] = build_cleanup_excel_bytes(filename_results, ocr_results)
    except Exception as exc:
        st.session_state["intelligent_cleanup_error"] = str(exc)


def build_cleanup_excel_bytes(filename_results: list[dict], ocr_results: list[dict]) -> bytes:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "识别清洗结果"
    headers = [
        "来源",
        "原始文件",
        "样本号",
        "订单编号",
        "物料编码",
        "物料ID",
        "成本中心",
        "报表类型",
        "证据类型",
        "建议标准文件名",
        "置信度",
        "OCR行数",
        "备注",
    ]
    _write_header(summary_sheet, headers)
    row_index = 2
    for item in filename_results:
        _write_cleanup_row(summary_sheet, row_index, "DeepSeek 文件名清洗", item)
        row_index += 1
    for item in ocr_results:
        cleaned = item.get("cleaned") or {}
        _write_cleanup_row(
            summary_sheet,
            row_index,
            "Qwen-OCR + DeepSeek",
            cleaned,
            source_file=item.get("source_file"),
            ocr_line_count=item.get("ocr_line_count"),
        )
        row_index += 1
    summary_sheet.freeze_panes = "A2"
    for column in summary_sheet.columns:
        summary_sheet.column_dimensions[column[0].column_letter].width = min(
            max(len(str(cell.value or "")) for cell in column) + 2,
            42,
        )

    ocr_sheet = workbook.create_sheet("OCR原文")
    _write_header(ocr_sheet, ["原始文件", "OCR原文"])
    for row_index, item in enumerate(ocr_results, start=2):
        ocr_sheet.cell(row=row_index, column=1, value=item.get("source_file"))
        ocr_sheet.cell(row=row_index, column=2, value=item.get("ocr_text"))
    ocr_sheet.column_dimensions["A"].width = 42
    ocr_sheet.column_dimensions["B"].width = 100
    ocr_sheet.freeze_panes = "A2"

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_filename_cleanup_excel_bytes(cleanup_results: list[dict]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "标准文件名清洗结果"
    headers = [
        "上传位置",
        "原始文件",
        "样本号",
        "订单编号",
        "物料编码",
        "物料ID",
        "成本中心",
        "识别类型",
        "证据类型",
        "建议标准文件名",
        "字段检查",
        "缺失字段",
    ]
    _write_header(sheet, headers)
    for row_index, item in enumerate(cleanup_results, start=2):
        values = [
            item.get("上传位置"),
            item.get("source_file"),
            item.get("sample_no"),
            item.get("order_id"),
            item.get("product_id"),
            item.get("material_id"),
            item.get("cost_center"),
            item.get("report_type"),
            item.get("evidence_kind"),
            item.get("standard_filename"),
            item.get("field_status"),
            "、".join(item.get("missing_field_labels") or []),
        ]
        for column_index, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=column_index, value=value)
    sheet.freeze_panes = "A2"
    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = min(
            max(len(str(cell.value or "")) for cell in column) + 2,
            52,
        )
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _excel_preview_rows(file_bytes: bytes, max_rows: int = 30) -> list[dict]:
    workbook = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = []
    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        values = list(row)
        if not any(value not in (None, "") for value in values):
            continue
        rows.append({f"列{index + 1}": value for index, value in enumerate(values)})
        if len(rows) >= max_rows:
            break
    workbook.close()
    return rows


def render_cleanup_file_preview(item: dict) -> None:
    file_name = item.get("standard_filename") or item.get("source_file") or "文件预览"
    file_bytes = item.get("file_bytes") or b""
    suffix = Path(str(file_name)).suffix.lower()
    st.caption(item.get("source_file") or "")
    if suffix in {".png", ".jpg", ".jpeg"}:
        st.image(file_bytes, caption=file_name, use_container_width=True)
    elif suffix in {".xlsx", ".xlsm"}:
        rows = _excel_preview_rows(file_bytes)
        if rows:
            st.dataframe(rows, use_container_width=True, hide_index=True)
        else:
            st.info("该 Excel 文件未读取到可预览数据。")
    elif suffix == ".csv":
        text = file_bytes.decode("utf-8-sig", errors="replace")
        st.text("\n".join(text.splitlines()[:40]))
    elif suffix == ".pdf":
        st.info("PDF 文件暂不支持在线预览，请下载标准命名文件包后查看。")
    else:
        st.info("该文件类型暂不支持在线预览。")


if hasattr(st, "dialog"):
    @st.dialog("文件预览")
    def show_cleanup_preview_dialog(item: dict) -> None:
        render_cleanup_file_preview(item)
        if st.button("关闭预览"):
            st.session_state.pop("cleanup_preview_index", None)
            st.rerun()


def _write_header(sheet, headers: list[str]) -> None:
    fill = PatternFill("solid", fgColor="00338D")
    font = Font(color="FFFFFF", bold=True)
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column_index, value=header)
        cell.fill = fill
        cell.font = font


def _write_cleanup_row(
    sheet,
    row_index: int,
    source: str,
    item: dict,
    *,
    source_file: str | None = None,
    ocr_line_count: int | None = None,
) -> None:
    values = [
        source,
        source_file or item.get("source_file"),
        item.get("sample_no"),
            item.get("order_id"),
            item.get("product_id"),
            item.get("material_id"),
            item.get("cost_center"),
            item.get("report_type"),
        item.get("evidence_kind"),
        item.get("standard_filename"),
        item.get("confidence"),
        ocr_line_count,
        item.get("notes"),
    ]
    for column_index, value in enumerate(values, start=1):
        sheet.cell(row=row_index, column=column_index, value=value)


deepseek_config = load_deepseek_config(st.secrets)
online_ocr_config = load_online_ocr_config(st.secrets)


st.markdown(
    f"""
    <a class="theme-button" href="?theme={next_theme}" target="_self" rel="noopener">{theme_label}</a>
    <a class="fork-button" href="https://github.com/tintindd/ai-mesp-workpaper-copilot" target="_blank" rel="noopener">
      <span>Fork</span>
      <svg viewBox="0 0 16 16" aria-hidden="true">
        <path d="M8 0C3.58 0 0 3.64 0 8.13c0 3.59 2.29 6.63 5.47 7.7.4.08.55-.18.55-.39 0-.19-.01-.83-.01-1.51-2.01.38-2.53-.5-2.69-.96-.09-.24-.48-.96-.82-1.15-.28-.15-.68-.53-.01-.54.63-.01 1.08.59 1.23.83.72 1.23 1.87.88 2.33.67.07-.53.28-.88.51-1.08-1.78-.2-3.64-.91-3.64-4.03 0-.89.31-1.62.82-2.19-.08-.2-.36-1.04.08-2.16 0 0 .67-.22 2.2.84A7.5 7.5 0 0 1 8 3.89c.68 0 1.36.09 2 .27 1.52-1.06 2.19-.84 2.19-.84.44 1.12.16 1.96.08 2.16.51.57.82 1.3.82 2.19 0 3.13-1.87 3.82-3.65 4.03.29.25.54.75.54 1.52 0 1.1-.01 1.99-.01 2.26 0 .22.15.47.55.39A8.08 8.08 0 0 0 16 8.13C16 3.64 12.42 0 8 0Z"/>
      </svg>
    </a>
    <div class="mesp-hero">
      <div>
        <h1>AI-MESP Workpaper Copilot</h1>
        <p>面向 MESP 底稿的智能取数、自动重算与异常复核助手</p>
      </div>
      <div class="mesp-badge">Local Demo · SAP Evidence Mapping</div>
    </div>
    """,
    unsafe_allow_html=True,
)

st.markdown('<div class="mesp-shell">', unsafe_allow_html=True)
step_col, work_col = st.columns([1.05, 3.75], gap="large")
active_step = get_active_step()
unlocked_step = get_unlocked_step()

with step_col:
    with st.container(border=True):
        nav_items = [
            (1, "1  选择场景", "成本方法、程序和审计期间"),
            (2, "2  输入参数", "样本文件与命名规则"),
            (3, "3  证据处理与上传", "OCR、清洗和计算结果"),
            (4, "4  复核结果", "异常、映射和追溯链"),
        ]
        for step, label, caption in nav_items:
            st.button(
                label,
                key=f"nav_step_{step}",
                type="primary" if step == active_step else "secondary",
                disabled=step > unlocked_step,
                use_container_width=True,
                on_click=go_to_step,
                args=(step,),
            )
            st.caption(caption if step <= unlocked_step else f"{caption}（未解锁）")

with work_col:
    if active_step == 1:
        with st.container(border=True):
            st.markdown(
                """
                <div class="main-title">
                  <h2>选择测试场景</h2>
                  <p>先选择成本方法、审计期间和 MESP 测试程序。系统会据此判断需要检查的支持文件和字段。</p>
                </div>
                """,
                unsafe_allow_html=True,
            )

            method_col, period_col = st.columns([1.1, 1.1], gap="large")
            with method_col:
                st.selectbox(
                    "成本方法",
                    ["标准成本法"],
                    index=0,
                    key="cost_method",
                    on_change=mark_scenario_started,
                )
            with period_col:
                st.text_input(
                    "审计期间",
                    value="2025.01.01-2025.12.31",
                    key="audit_period",
                    on_change=mark_scenario_started,
                )

            st.markdown(
                """
                <div class="program-help">
                  <div class="program-card">
                    <strong>标准成本法 - 固定加工成本</strong>
                    <div class="program-code">SPD03012</div>
                    <span class="pill">CO03</span><span class="pill">KSBT</span><span class="pill">3611</span>
                  </div>
                  <div class="program-card">
                    <strong>标准成本法 - 可变加工成本</strong>
                    <div class="program-code">SPD03014</div>
                    <span class="pill">CO03</span><span class="pill">KSBT</span><span class="pill">3611</span>
                  </div>
                  <div class="program-card">
                    <strong>存货成本差异分摊</strong>
                    <div class="program-code">SPD03015</div>
                    <span class="pill">CKM3</span><span class="pill">3611</span>
                  </div>
                </div>
                """,
                unsafe_allow_html=True,
            )

            current_program = st.session_state.get("test_program", "SPD03012")
            default_label = next(
                (label for label, code in PROGRAM_OPTIONS.items() if code == current_program),
                "SPD03012 - 标准成本法固定加工成本",
            )
            st.selectbox(
                "测试程序",
                list(PROGRAM_OPTIONS),
                index=list(PROGRAM_OPTIONS).index(default_label),
                key="test_program_label",
                on_change=sync_test_program_from_label,
            )
            st.session_state["test_program"] = PROGRAM_OPTIONS.get(
                st.session_state.get("test_program_label"),
                current_program,
            )

            st.button("下一步", type="primary", on_click=go_next_from_scenario)

    elif active_step == 2:
        with st.container(border=True):
            st.markdown(
                """
                <div class="main-title">
                  <h2>输入参数与命名规则</h2>
                  <p>确认审计期间、测试程序和样本文件命名要求。确认后才会解锁上传证据。</p>
                </div>
                """,
                unsafe_allow_html=True,
            )

            summary_cols = st.columns(3)
            summary_cols[0].metric("成本方法", st.session_state.get("cost_method", "标准成本法"))
            summary_cols[1].metric("测试程序", st.session_state.get("test_program", "SPD03012"))
            summary_cols[2].metric("审计期间", st.session_state.get("audit_period", "2025.01.01-2025.12.31"))

            current_program = st.session_state.get("test_program", "SPD03012")
            required_reports = required_reports_for_program(current_program)
            requires_order_id = program_requires_order_id(current_program)
            required_examples = []
            if "CO03" in required_reports:
                required_examples.append("<code>样本3/3.订单编号11001846-CO03-表格.xlsx</code>")
            if "KSBT" in required_reports:
                required_examples.append("<code>样本3/3.订单编号11001846-KSBT-表格.xlsx</code>")
            if "3611" in required_reports:
                required_examples.append(
                    "<code>样本3/3.订单编号11001846-3611-表格.xlsx</code>"
                    if requires_order_id
                    else "<code>样本3/3.3611-表格.xlsx</code>"
                )
            if "CKM3" in required_reports:
                required_examples.append(
                    "<code>样本3/3.订单编号11001846-物料ID-13012857-CKM3-表格.xlsx</code>"
                    if requires_order_id
                    else "<code>样本3/3.物料ID-13012857-CKM3-表格.xlsx</code>"
                )
            filename_fields = "样本号、订单编号和报表类型" if requires_order_id else "样本号、报表类型；CKM3 还需包含物料ID"
            st.markdown(
                f"""
                <div class="upload-rules">
                  <strong>命名要求</strong><br>
                  zip 包内建议按 <code>样本1/</code>、<code>样本2/</code> 建文件夹；每个样本至少包含
                  <code>{"</code>、<code>".join(required_reports)}</code> 表格文件，可同时包含对应截图。<br>
                  文件名需包含{filename_fields}，例如：
                  {"、".join(required_examples)}。
                </div>
                """,
                unsafe_allow_html=True,
            )

            back_col, next_col = st.columns([1, 1])
            with back_col:
                st.button("返回上一步", on_click=go_to_step, args=(1,))
            with next_col:
                st.button("确认参数，下一步", type="primary", on_click=go_next_from_params)

    elif active_step == 3:
        with st.container(border=True):
            st.markdown(
                """
                <div class="main-title">
                  <h2>证据处理与上传</h2>
                  <p>在同一个工作区完成可选的 OCR、文件名标准化清洗，以及计算结果生成。若已准备好标准文件，可直接使用计算结果。</p>
                </div>
                """,
                unsafe_allow_html=True,
            )

            active_program = st.session_state.get("test_program", "SPD03012")
            required_reports = required_reports_for_program(active_program)
            requires_co03 = "CO03" in required_reports
            requires_ksbt = "KSBT" in required_reports
            requires_ckm3 = "CKM3" in required_reports
            requires_order_id = program_requires_order_id(active_program)
            first_step_text = (
                '1. 若缺少 CKM3 Excel，先在 <code>CKM3表格OCR</code> 中把截图转成表格。<br>'
                if requires_ckm3
                else "1. 若文件命名不规范，直接进入文件名清洗。<br>"
            )
            st.markdown(
                f"""
                <div class="upload-rules">
                  <strong>推荐处理顺序</strong><br>
                  {first_step_text}
                  2. 若文件命名不规范，在 <code>文件名清洗</code> 中导出标准命名 ZIP。<br>
                  3. 将最终标准文件或 zip 包上传到 <code>计算结果</code>，生成 SPP 和底稿结果。
                </div>
                """,
                unsafe_allow_html=True,
            )
            st.info(
                f"当前选择的测试程序：{active_program}；本步骤仅需要："
                + "、".join(required_reports)
            )

            st.markdown("### 1. 参数获取")
            param_mode = st.radio(
                "参数来源",
                ["OCR识别", "手工输入"],
                horizontal=True,
                key="param_capture_mode",
            )
            params = current_evidence_params()
            if param_mode == "手工输入":
                manual_fields = [
                    ("sample_no", "样本编号", params.get("sample_no", "1"), "manual_sample_no"),
                ]
                if requires_order_id:
                    manual_fields.append(("order_id", "订单编号", params.get("order_id", ""), "manual_order_id"))
                if requires_co03:
                    manual_fields.append(("product_id", "CO03物料编码", params.get("product_id", ""), "manual_product_id"))
                if requires_ckm3:
                    manual_fields.append(("material_id", "CKM3物料ID", params.get("material_id", ""), "manual_material_id"))
                if requires_ksbt:
                    manual_fields.append(("cost_center", "KSBT成本中心", params.get("cost_center", ""), "manual_cost_center"))
                manual_values = {
                    "sample_no": params.get("sample_no", "1"),
                    "order_id": params.get("order_id", ""),
                    "product_id": params.get("product_id", ""),
                    "material_id": params.get("material_id", ""),
                    "cost_center": params.get("cost_center", ""),
                }
                manual_cols = st.columns(len(manual_fields))
                for column, (field, label, value, key) in zip(manual_cols, manual_fields):
                    manual_values[field] = column.text_input(label, value=value, key=key)
                st.button(
                    "确认手工参数",
                    type="primary",
                    on_click=apply_manual_params,
                    args=(
                        manual_values["sample_no"],
                        manual_values["order_id"],
                        manual_values["product_id"],
                        manual_values["material_id"],
                        manual_values["cost_center"],
                    ),
                )
            else:
                ocr_requirements = []
                if requires_co03:
                    ocr_requirements.append("CO03截图识别订单编号和物料编码")
                if requires_ksbt:
                    ocr_requirements.append("KSBT截图识别成本中心")
                if requires_ckm3:
                    ocr_requirements.append("CKM3截图识别物料ID")
                st.caption("；".join(ocr_requirements) + "。")
                ocr_field_count = 1 + int(requires_co03) + int(requires_ksbt) + int(requires_ckm3)
                ocr_cols = st.columns(ocr_field_count)
                ocr_column_index = 0
                ocr_sample_no = ocr_cols[ocr_column_index].text_input("样本编号", value=params.get("sample_no", "1"), key="ocr_sample_no")
                ocr_column_index += 1
                co03_param_file = None
                ksbt_param_file = None
                ckm3_param_file = None
                if requires_co03:
                    co03_param_file = ocr_cols[ocr_column_index].file_uploader("CO03截图", type=["png", "jpg", "jpeg"], key="param_co03_file")
                    ocr_column_index += 1
                if requires_ksbt:
                    ksbt_param_file = ocr_cols[ocr_column_index].file_uploader("KSBT截图", type=["png", "jpg", "jpeg"], key="param_ksbt_file")
                    ocr_column_index += 1
                if requires_ckm3:
                    ckm3_param_file = ocr_cols[ocr_column_index].file_uploader("CKM3截图", type=["png", "jpg", "jpeg"], key="param_ckm3_file")
                st.button(
                    "开始OCR识别参数",
                    disabled=not online_ocr_available(online_ocr_config) or not any([co03_param_file, ksbt_param_file, ckm3_param_file]),
                    type="primary",
                    on_click=run_parameter_ocr,
                    args=(ocr_sample_no, co03_param_file, ckm3_param_file, ksbt_param_file),
                )
                if st.session_state.get("parameter_ocr_error"):
                    st.error(st.session_state["parameter_ocr_error"])

            st.markdown("#### 参数识别台账")
            evidence_editor_rows = [
                {"_selected_for_delete": False, **row}
                for row in evidence_params_table()
            ]
            parameter_column_order = ["_selected_for_delete", "sample_no"]
            if requires_order_id:
                parameter_column_order.append("order_id")
            if requires_co03:
                parameter_column_order.append("product_id")
            if requires_ckm3:
                parameter_column_order.append("material_id")
            if requires_ksbt:
                parameter_column_order.append("cost_center")
            edited_params = st.data_editor(
                evidence_editor_rows,
                use_container_width=True,
                hide_index=True,
                key="evidence_params_editor",
                column_order=parameter_column_order,
                column_config={
                    "_selected_for_delete": st.column_config.CheckboxColumn("", width="small"),
                    "sample_no": st.column_config.TextColumn("样本编号", required=True),
                    "order_id": st.column_config.TextColumn("订单编号"),
                    "product_id": st.column_config.TextColumn("CO03物料编码"),
                    "material_id": st.column_config.TextColumn("CKM3物料ID"),
                    "cost_center": st.column_config.TextColumn("KSBT成本中心"),
                },
            )
            apply_evidence_params_editor(edited_params or [])
            selected_samples_to_delete = [
                str(row.get("sample_no") or "").strip()
                for row in (edited_params or [])
                if row.get("_selected_for_delete") and str(row.get("sample_no") or "").strip()
            ]
            if selected_samples_to_delete:
                st.warning(
                    "已选择删除样本："
                    + "、".join(f"样本{sample}" for sample in selected_samples_to_delete)
                    + "。点击下方删除按钮后才会删除，并同步清理这些样本的已清洗/补充文件。"
                )
                if st.button("删除", type="secondary"):
                    delete_evidence_samples(set(selected_samples_to_delete))
                    st.rerun()
            params = current_evidence_params()

            preview_items = st.session_state.get("parameter_ocr_previews") or []
            if preview_items:
                with st.expander("查看参数OCR截图预览", expanded=False):
                    for item in preview_items:
                        st.markdown(f"**{item.get('label')}：{item.get('source_file')}**")
                        st.image(item.get("file_bytes"), use_container_width=True)
                keep_for_cleanup = st.checkbox("确认并保留这些截图到下一步文件名清洗", value=True, key="keep_param_files_for_cleanup")
                confirm_cols = st.columns([1, 1, 3])
                confirm_cols[0].button(
                    "确认参数和截图",
                    type="primary",
                    on_click=confirm_parameter_standard_files,
                    args=(not keep_for_cleanup,),
                )
                if st.session_state.get("parameter_standard_files"):
                    param_zip = build_standard_named_zip_bytes(st.session_state["parameter_standard_files"])
                    confirm_cols[1].download_button(
                        "下载参数截图ZIP",
                        data=param_zip,
                        file_name=f"样本{params.get('sample_no') or '1'}_参数截图标准命名.zip",
                        mime="application/zip",
                        type="secondary",
                    )

            st.markdown("### 2. 文件名清洗与计算")
            if requires_ckm3:
                tab_ocr, tab_cleanup, tab_upload = st.tabs(["CKM3表格OCR", "文件名清洗", "计算结果"])
                with tab_ocr:
                    ocr_col, config_col = st.columns([1.25, 1])
                    with ocr_col:
                        st.info("仅用于 CKM3 截图，识别后生成可替代 CKM3-表格.xlsx 的支持性 Excel。3611 仍优先使用 SAP 导出的 Excel/CSV。")
                    with config_col:
                        if online_ocr_available(online_ocr_config):
                            st.success("Qwen-OCR 服务已配置，智谱视觉可作为备用通道")
                        else:
                            st.warning("Qwen-OCR 服务未配置，OCR 暂不可用。")

                    ckm3_ocr_files = st.file_uploader(
                        "上传 CKM3 截图（PNG/JPG/JPEG/PDF）",
                        type=["png", "jpg", "jpeg", "pdf"],
                        accept_multiple_files=True,
                        key="ckm3_ocr_files",
                    )
                    st.button(
                        "仅识别 CKM3 截图并生成 Excel",
                        disabled=not ckm3_ocr_files or not online_ocr_available(online_ocr_config),
                        type="primary",
                        on_click=run_ckm3_ocr_to_excel,
                        args=(ckm3_ocr_files,),
                    )

                    ckm3_ocr_error = st.session_state.get("ckm3_ocr_error")
                    if ckm3_ocr_error:
                        st.error(ckm3_ocr_error)

                    ckm3_ocr_excels = st.session_state.get("ckm3_ocr_excels") or []
                    if ckm3_ocr_excels:
                        st.dataframe(
                            [
                                {
                                    "来源截图": item.get("source_file"),
                                    "状态": item.get("status"),
                                    "生成文件": item.get("file_name") or "-",
                                    "明细行数": item.get("row_count"),
                                }
                                for item in ckm3_ocr_excels
                            ],
                            use_container_width=True,
                            hide_index=True,
                        )
                        for index, item in enumerate(ckm3_ocr_excels, start=1):
                            if item.get("bytes"):
                                st.download_button(
                                    f"下载 OCR 生成 Excel #{index}",
                                    data=item["bytes"],
                                    file_name=Path(item.get("file_name") or f"ckm3_ocr_{index}.xlsx").name,
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    type="secondary",
                                )
                        with st.expander("查看 OCR 原文"):
                            for item in ckm3_ocr_excels:
                                st.markdown(f"**{item.get('source_file')}**")
                                st.text(item.get("ocr_text") or "")
            else:
                tab_cleanup, tab_upload = st.tabs(["文件名清洗", "计算结果"])

            with tab_cleanup:
                config_col, action_col = st.columns([1.2, 1])
                with config_col:
                    if deepseek_config:
                        st.success(f"DeepSeek 已配置：{deepseek_config.model}")
                    else:
                        st.warning("DeepSeek API Key 未配置，将仅按输入信息和上传位置生成标准文件名。")
                        st.caption("在 Streamlit Cloud 的 Secrets 中添加 DEEPSEEK_API_KEY 后，可辅助识别原始文件名。")
                with action_col:
                    st.button(
                        "测试 DeepSeek 连接",
                        disabled=not deepseek_config,
                        on_click=run_deepseek_connection_test,
                    )
                    deepseek_status = st.session_state.get("deepseek_status")
                    if deepseek_status:
                        if deepseek_status.get("ok"):
                            st.caption(deepseek_status.get("message"))
                        else:
                            st.error(deepseek_status.get("message"))

                params = current_evidence_params()
                st.session_state["cleanup_sample_count"] = evidence_sample_count()
                sample_total_col, sample_col = st.columns(2)
                with sample_total_col:
                    cleanup_sample_count = st.number_input(
                        "样本总数",
                        min_value=1,
                        max_value=50,
                        value=evidence_sample_count(),
                        step=1,
                        key="cleanup_sample_count",
                        disabled=True,
                    )
                with sample_col:
                    cleanup_sample_no = st.selectbox(
                        "当前清洗样本",
                        [str(index) for index in range(1, int(cleanup_sample_count) + 1)],
                        index=max(min(int(params.get("sample_no") or 1), int(cleanup_sample_count)) - 1, 0),
                        key="cleanup_sample_no_select",
                    )
                params = current_evidence_params(cleanup_sample_no)
                cleanup_order_id = params.get("order_id") or "" if requires_order_id else ""
                st.markdown("##### 样本文件完整性")
                completeness = sample_completeness_rows_v2(
                    int(cleanup_sample_count),
                    st.session_state.get("filename_cleanup_results") or [],
                    required_reports,
                )
                st.dataframe(completeness, use_container_width=True, hide_index=True)
                example_names = []
                if "CO03" in required_reports:
                    example_names.append("<code>样本1/1.订单编号11000437-CO03-表格.xlsx</code>")
                if "KSBT" in required_reports:
                    example_names.append("<code>样本1/1.订单编号11000437-KSBT-截图.png</code>")
                if "3611" in required_reports:
                    example_names.append(
                        "<code>样本1/1.订单编号11000437-3611-截图.png</code>"
                        if requires_order_id
                        else "<code>样本1/1.3611-截图.png</code>"
                    )
                if "CKM3" in required_reports:
                    example_names.append(
                        "<code>样本1/1.订单编号11000437-物料ID-13014012-CKM3-截图.png</code>"
                        if requires_order_id
                        else "<code>样本1/1.物料ID-13014012-CKM3-截图.png</code>"
                    )
                st.markdown(
                    """
                    <div class="upload-rules">
                      <strong>文件名清洗说明</strong><br>
                      当前测试程序仅需要 <code>{required_reports_text}</code>。请按文件所属类型上传到对应位置。输出文件包将按标准格式重命名，例如
                      {example_text}。
                      只会对当前程序所需资料生成标准命名文件。
                    </div>
                    """.format(
                        required_reports_text="、".join(required_reports),
                        example_text="、".join(example_names),
                    ),
                    unsafe_allow_html=True,
                )

                bucket_cols = st.columns(len(required_reports))
                bucket_files = {}
                for column, report in zip(bucket_cols, required_reports):
                    with column:
                        bucket_files[report] = st.file_uploader(
                            f"{report} 文件",
                            type=["xlsx", "xlsm", "csv", "png", "jpg", "jpeg", "pdf"],
                            accept_multiple_files=True,
                            key=f"cleanup_{report.lower()}_files",
                        )

                total_cleanup_files = sum(len(files or []) for files in bucket_files.values())
                param_status = {}
                if requires_order_id:
                    param_status["订单编号"] = params.get("order_id") or "未填写"
                if requires_co03:
                    param_status["CO03物料编码"] = params.get("product_id") or "未填写"
                if requires_ckm3:
                    param_status["CKM3物料ID"] = params.get("material_id") or "未填写"
                if requires_ksbt:
                    param_status["KSBT成本中心"] = params.get("cost_center") or "未填写"
                st.caption(
                    "当前样本关键参数："
                    + "；".join(f"{key}={value}" for key, value in param_status.items())
                )
                st.button(
                    "运行文件名清洗",
                    disabled=total_cleanup_files == 0,
                    type="primary",
                    on_click=run_filename_cleanup_by_bucket,
                    args=(cleanup_sample_no, cleanup_order_id, bucket_files),
                )

                cleanup_error = st.session_state.get("filename_cleanup_error")
                if cleanup_error:
                    st.error(cleanup_error)

                cleanup_results = st.session_state.get("filename_cleanup_results") or []
                if cleanup_results:
                    cleanup_table_rows = []
                    for item in cleanup_results:
                        cleanup_row = {
                            "上传位置": item.get("上传位置"),
                            "原始文件": item.get("source_file"),
                            "样本号": item.get("sample_no") or "-",
                        }
                        if requires_order_id:
                            cleanup_row["订单编号"] = item.get("order_id") or "-"
                        if requires_co03:
                            cleanup_row["物料编码"] = item.get("product_id") or "-"
                        if requires_ckm3:
                            cleanup_row["物料ID"] = item.get("material_id") or "-"
                        if requires_ksbt:
                            cleanup_row["成本中心"] = item.get("cost_center") or "-"
                        cleanup_row.update(
                            {
                                "识别类型": item.get("report_type") or "-",
                                "证据类型": item.get("evidence_kind") or "-",
                                "建议标准文件名": item.get("standard_filename") or "-",
                                "字段检查": item.get("field_status") or "-",
                                "缺失字段": "、".join(item.get("missing_field_labels") or []) or "-",
                            }
                        )
                        cleanup_table_rows.append(cleanup_row)
                    st.dataframe(cleanup_table_rows, use_container_width=True, hide_index=True)

                    st.markdown("#### 标准文件名预览")
                    st.caption("点击标准文件名可预览文件内容。图片会直接展示，Excel 会展示首个 Sheet 的前 30 行。")
                    for index, item in enumerate(cleanup_results):
                        row_cols = st.columns([1.1, 3.2, 1.2, 1])
                        row_cols[0].write(item.get("上传位置") or "-")
                        row_cols[1].button(
                            item.get("standard_filename") or item.get("source_file") or f"文件 {index + 1}",
                            key=f"preview_cleanup_{index}",
                            use_container_width=True,
                            on_click=lambda i=index: st.session_state.update({"cleanup_preview_index": i}),
                        )
                        row_cols[2].write(item.get("evidence_kind") or "-")
                        row_cols[3].write(item.get("field_status") or "-")

                    preview_index = st.session_state.get("cleanup_preview_index")
                    if preview_index is not None and preview_index < len(cleanup_results):
                        preview_item = cleanup_results[preview_index]
                        if hasattr(st, "dialog"):
                            show_cleanup_preview_dialog(preview_item)
                        else:
                            with st.expander("文件预览", expanded=True):
                                if st.button("关闭预览"):
                                    st.session_state.pop("cleanup_preview_index", None)
                                    st.rerun()
                                render_cleanup_file_preview(preview_item)

                    standard_zip_bytes = st.session_state.get("standard_named_zip_bytes")
                    if standard_zip_bytes:
                        st.download_button(
                            "导出标准命名文件包 ZIP",
                            data=standard_zip_bytes,
                            file_name=f"样本{cleanup_sample_no or '1'}_标准命名文件包.zip",
                            mime="application/zip",
                            type="primary",
                        )
                    calculate_choice = st.radio(
                        "是否进行计算",
                        ["否", "是"],
                        horizontal=True,
                        key="calculate_after_cleanup",
                    )
                    if calculate_choice == "是":
                        st.button(
                            "使用清洗后的标准文件直接计算",
                            disabled=not standard_zip_bytes,
                            type="primary",
                            on_click=run_analysis_from_cleaned_zip,
                            args=(
                                standard_zip_bytes,
                                st.session_state.get("audit_period", "2025.01.01-2025.12.31"),
                                st.session_state.get("test_program", "SPD03012"),
                            ),
                        )
                    st.download_button(
                        "下载标准文件名清洗清单 Excel",
                        data=build_filename_cleanup_excel_bytes(cleanup_results),
                        file_name="AI-MESP_标准文件名清洗结果.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="secondary",
                    )

            with tab_upload:
                calculation_sample_count = evidence_sample_count()
                st.markdown("##### 样本文件完整性")
                st.dataframe(
                    sample_completeness_rows_v2(
                        int(calculation_sample_count),
                        st.session_state.get("filename_cleanup_results") or [],
                        required_reports,
                    ),
                    use_container_width=True,
                    hide_index=True,
                )

                st.markdown("##### 补充缺失文件")
                missing_cols = st.columns(4)
                with missing_cols[0]:
                    missing_sample_no = st.selectbox(
                        "补充样本",
                        [str(index) for index in range(1, int(calculation_sample_count) + 1)],
                        key="missing_file_sample_no",
                    )
                with missing_cols[1]:
                    missing_report = st.selectbox(
                        "文件类型",
                        required_reports,
                        key="missing_file_report",
                    )
                with missing_cols[2]:
                    missing_kind = st.selectbox(
                        "证据类型",
                        ["表格", "截图"],
                        key="missing_file_kind",
                    )
                with missing_cols[3]:
                    missing_file = st.file_uploader(
                        "上传缺失文件",
                        type=["xlsx", "xlsm", "csv", "png", "jpg", "jpeg", "pdf"],
                        key="missing_support_file",
                    )
                st.button(
                    "添加到样本文件池",
                    disabled=not missing_file,
                    type="secondary",
                    on_click=add_missing_support_file,
                    args=(missing_sample_no, missing_report, missing_kind, missing_file),
                )
                if st.session_state.get("missing_support_file_error"):
                    st.error(st.session_state["missing_support_file_error"])

                standard_zip_bytes = st.session_state.get("standard_named_zip_bytes")
                if standard_zip_bytes:
                    st.button(
                        "使用已清洗/补充文件计算",
                        type="primary",
                        on_click=run_analysis_from_cleaned_zip,
                        args=(
                            standard_zip_bytes,
                            st.session_state.get("audit_period", "2025.01.01-2025.12.31"),
                            st.session_state.get("test_program", "SPD03012"),
                        ),
                    )

                st.markdown("##### 直接上传标准文件计算")
                uploaded_files = st.file_uploader(
                    "上传最终支持文件或 zip 包并计算",
                    type=["xlsx", "xlsm", "csv", "png", "jpg", "jpeg", "pdf", "zip"],
                    accept_multiple_files=True,
                    key="supporting_files",
                )

                if uploaded_files:
                    st.session_state["has_uploaded_files"] = True
                    st.write(f"已选择 {len(uploaded_files)} 个文件。")

                analyze_clicked = st.button("Analyze", type="primary", disabled=not uploaded_files)

        if analyze_clicked:
            period = st.session_state.get("audit_period", "2025.01.01-2025.12.31")
            program = st.session_state.get("test_program", "SPD03012")
            with tempfile.TemporaryDirectory(prefix="mesp_streamlit_") as temp:
                temp_dir = Path(temp)
                saved_count = save_uploaded_files(uploaded_files, temp_dir)
                if saved_count == 0:
                    st.error("没有可分析的文件。")
                    st.stop()

                with st.spinner("正在识别 SAP 支持文件并生成复核结果..."):
                    st.session_state["analysis_bundle"] = build_analysis_bundle_from_folder(temp_dir, period, program)
            st.session_state["current_step"] = 4
            st.rerun()

    elif active_step == 4:
        bundle = st.session_state.get("analysis_bundle")
        if not bundle:
            st.info("请先上传证据并点击 Analyze，完成后这里会展示复核结果。")
        else:
            result = bundle["result"]
            selected_program = bundle["program"]
            supporting_bytes = bundle["supporting_bytes"]
            selected_spd_bytes = bundle["selected_spd_bytes"]

            summary = result.get("summary", {})
            cols = st.columns(5)
            cols[0].metric("样本数", summary.get("sample_count", 0))
            cols[1].metric("识别文件", summary.get("recognized_file_count", 0))
            cols[2].metric("异常项", summary.get("issue_count", len(result.get("issues") or [])))
            cols[3].metric("表格数", summary.get("workbook_count", 0))
            cols[4].metric("已追溯项", summary.get("trace_count", len(result.get("evidence_trace") or [])))

            tab_issues, tab_workbooks, tab_trace, tab_download = st.tabs(
                ["异常与追问", "SAP 表格映射", "Evidence Traceability", "底稿结果下载"]
            )

            with tab_issues:
                render_issues(result.get("issues") or [])

            with tab_workbooks:
                render_workbooks(result.get("workbook_results") or result.get("co03_results") or [])

            with tab_trace:
                render_trace(result.get("evidence_trace") or [])

            with tab_download:
                if summary.get("sample_count", 0) == 0 or summary.get("recognized_file_count", 0) == 0:
                    st.warning("未识别到有效样本，暂不生成底稿结果。请先根据异常提示修正上传文件命名或内容。")
                else:
                    st.info(f"本次生成的测试程序：{selected_program}")
                    st.download_button(
                        "下载 SPP Supporting Excel",
                        data=supporting_bytes,
                        file_name="AI-MESP_SPP_Supporting.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="secondary",
                    )
                    st.download_button(
                        f"下载 {selected_program}_IRM(SAP)",
                        data=selected_spd_bytes,
                        file_name=f"AI-MESP_{selected_program}_IRM(SAP).xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                    )

st.markdown("</div>", unsafe_allow_html=True)
